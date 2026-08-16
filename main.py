import argparse
import asyncio
import datetime
import functools
import hashlib
import io
import itertools
import lzma
import os
import re
import shutil
import sqlite3
import stat as statmod
import sys
import time
from argparse import Namespace
from collections import Counter, defaultdict, deque
from collections.abc import AsyncGenerator, Callable, Generator
from contextlib import asynccontextmanager, contextmanager
from itertools import combinations
from pathlib import Path
from typing import Any, Final, NamedTuple
from urllib.parse import urlparse

import aiosqlite
import fasteners
import numpy as np
from mtproxy_bridge import is_mtproto_link, needs_padded_transport, start_local_bridge

# Используется kurigram
from pyrogram import Client, types
from pyrogram.connection.transport.tcp import TCPAbridged  # , TCPIntermediatePadded
from pyrogram.enums import ChatMemberStatus, ChatType, MessagesFilter
from pyrogram.errors import (
    PeerIdInvalid,
    UsernameInvalid,
    UsernameNotOccupied,
    UserNotParticipant,
)
from rapidfuzz import fuzz, process
from rapidfuzz.utils import default_process

from config import (
    _KEEP_CRITERIA_VALID,
    ABORT_DELETE_ON_ARCHIVE_FAILURE,
    API_HASH,
    API_ID,
    ARCHIVE_BEFORE_DELETE,
    ARCHIVE_HIDE_SENDER,
    ARCHIVE_MODE,
    ARCHIVE_OLD_BACKUPS,
    ARCHIVE_TARGET,
    BACKUP_DIR,
    BACKUP_ON_STARTUP,
    BACKUP_ONLY_IF_CHANGED,
    BATCH_DELETE_SIZE,
    CHAT_LABEL_PARTS,
    CHAT_LIST,
    DB_CACHE_SIZE,
    DB_FILE,
    DOWNLOADS_DIR,
    DRY_RUN,
    DURATION_POWER,
    DYNAMIC_SPACE_COEFFICIENT,
    DYNAMIC_SPACE_SAFETY_BUFFER_MB,
    ENABLE_FUZZY_MATCHING,
    EXPORTS_DIR,
    FUZZY_MATCHING_MODE,
    FUZZY_THRESHOLD,
    KEEP_PRIORITY,
    LOCK_TIMEOUT,
    LZMA_PRESET,
    MAX_ARCHIVES,
    MAX_BACKUPS,
    MAX_DURATION_DIFF_SEC,
    MIN_FREE_SPACE_MB,
    NAME_POWER,
    PENALTY_NUMBERS_MISMATCH,
    PROXY_URL,
    RAW_IGNORE_LIST,
    RAW_IGNORE_REGEX,
    REPORT_ONLY,
    REVOKE_PRIVATE_CHATS,
    ROTATE_BEFORE_BACKUP,
    SESSION_NAME,
    SIZE_POWER,
    SLEEP_THRESHOLD,
    SYNC_BATCH_SIZE,
    USE_JACCARD_PENALTY,
    USE_META_FUZZY,
    VERIFY_CHUNK_SIZE,
    VERIFY_CONCURRENCY,
    WEIGHT_DURATION,
    WEIGHT_NAME,
    WEIGHT_SIZE,
)
from logger import log
from tcp_padded import TCPPadded

LOCK_FILE = Path(f"{SESSION_NAME}.lock")

IGNORE_MESSAGES: Final[defaultdict[int, set[int]]] = defaultdict(set)
IGNORE_REGEX: Final[defaultdict[int, list[re.Pattern[str]]]] = defaultdict(list)
GLOBAL_IGNORE_REGEX: Final[list[re.Pattern[str]]] = []
CHAT_LABELS: dict[int, tuple[str, str | None]] = {}


# region === TYPE DEFINITIONS (Python 3.12+) ===

# Примитивы (для читаемости)
type ChatID = int
type MessageID = int
type FileUniqueID = str


# Структура, которую возвращает _get_audio_attributes
class AudioMeta(NamedTuple):
    """Атрибуты аудиосообщения (порядок полей = порядок колонок audios в БД)."""

    file_unique_id: FileUniqueID
    file_name: str | None
    file_size: int
    duration: int
    performer: str | None
    title: str | None


# Структура для одной строки из БД (обертка над sqlite Row)
type DBRow = aiosqlite.Row

# Группа дубликатов - это список строк из БД
type DuplicateGroup = list[DBRow]

# Алиас для функции форматирования строки CSV при экспорте
type CsvRowFormatter = Callable[[DBRow], list[str] | None]

# Ключ ребра графа дубликатов: упорядоченная пара (min message_id, max message_id)
type EdgeKey = tuple[MessageID, MessageID]


class EdgeInfo(NamedTuple):
    """Причина связи двух файлов и коэффициенты сходства.

    reason  — "uid" / "meta" / "fuzzy".
    score   — итоговый коэффициент сходства (для uid/meta = 1.0).
    name    — вклад текстового fuzzy (имя/мета), 0..1; None для uid/meta.
              (legacy-название поля; фактически это лучший текстовый источник)
    dur     — вклад длительности (0..1); None для uid/meta.
    size    — вклад размера (0..1); None для uid/meta.
    penalty — штраф за несовпадение числовых токенов (0.0 если нет).
    text_source — код источника fuzzy-совпадения (0..3); None для uid/meta.
    """

    reason: str
    score: float
    name: float | None
    dur: float | None
    size: float | None
    penalty: float
    text_source: int | None = None


# message_id-пара -> метаданные связи
type EdgeMeta = dict[EdgeKey, EdgeInfo]


def _edge_key(a: MessageID, b: MessageID) -> EdgeKey:
    """Канонический (неориентированный) ключ ребра."""
    return (a, b) if a < b else (b, a)


# Словари для верификации (ID сообщения -> Объект сообщения или Ошибка/None)
type VerifiedMessagesDict = dict[MessageID, types.Message | None | Exception]


# Результат классификации дубликатов
class ClassificationResult(NamedTuple):
    delete_from_tg: set[MessageID]  # удалить из Telegram
    delete_from_db: set[MessageID]  # удалить только из БД (сообщения нет в ТГ)
    update_in_db: list[types.Message]  # обновить в БД (контент изменился)


# Алиас для функции форматирования строки при экспорте
type RowFormatter = Callable[[DBRow], str | None]

# endregion

# region --- КЛАССЫ ОШИБОК ---


class AlreadyRunningError(RuntimeError):
    """Исключение, выбрасываемое при невозможности захватить lock-файл."""

    pass


class IgnoreListResolutionError(Exception):
    """Исключение, выбрасываемое, если не удалось разрешить идентификаторы из ignore_list."""

    pass


# endregion

# region --- СИСТЕМНЫЕ УТИЛИТЫ И ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ---
# Блок для общих, низкоуровневых задач, не связанных напрямую с бизнес-логикой: парсинг аргументов, блокировка, форматирование и работа с файловой системой.


def parse_arguments() -> argparse.Namespace:
    """Настраивает и парсит аргументы командной строки.

    Подкоманды: ``repair``, ``report``, ``download``, ``export``.
    Вызов без подкоманды — обычный прогон дедупликации.
    """
    parser = argparse.ArgumentParser(
        description="Скрипт для поиска и удаления дубликатов аудио в Telegram чатах."
    )
    subparsers = parser.add_subparsers(dest="command", metavar="<команда>")

    subparsers.add_parser(
        "repair",
        help="Запускает утилиту для восстановления и оптимизации базы данных.",
    )

    subparsers.add_parser(
        "report",
        help="Создает текстовый файл-отчет с найденными группами дубликатов и ссылками (без удаления).",
    )

    p_download = subparsers.add_parser(
        "download",
        help="Скачивает все аудиофайлы из БД для указанного чата в папку downloads.",
    )
    p_download.add_argument(
        "chat",
        type=str,
        metavar="CHAT_IDENTIFIER",
        help="Идентификатор чата (ID, @username или ссылка).",
    )

    p_export = subparsers.add_parser(
        "export",
        help="Экспорт данных из БД; действие указывается подкомандой.",
    )
    export_subparsers = p_export.add_subparsers(
        dest="export_command", metavar="<действие>", required=True
    )

    p = export_subparsers.add_parser(
        "filenames",
        help="Экспортирует все имена файлов из БД в текстовый файл и завершает работу.",
    )
    p.add_argument(
        "chat",
        type=chat_id_or_username,
        metavar="CHAT_ID|@USERNAME",
        help="Чат, для которого экспортировать имена файлов.",
    )

    p = export_subparsers.add_parser(
        "filenames-url",
        help="Экспортирует имена файлов и ссылки на сообщения из БД в текстовый файл.",
    )
    p.add_argument(
        "chat",
        type=chat_id_or_username,
        metavar="CHAT_ID|@USERNAME",
        help="Чат, для которого экспортировать имена файлов со ссылками.",
    )

    p = export_subparsers.add_parser(
        "cleaned-names",
        help="Экспортирует процесс очистки имён файлов в CSV формат.",
    )
    p.add_argument(
        "chat",
        type=chat_id_or_username,
        metavar="CHAT_ID|@USERNAME",
        nargs="?",
        default=0,
        help="Чат; без аргумента — вся БД.",
    )

    p = export_subparsers.add_parser(
        "cleaned-meta",
        help="Экспортирует процесс очистки метаданных (performer+title) в CSV формат.",
    )
    p.add_argument(
        "chat",
        type=chat_id_or_username,
        metavar="CHAT_ID|@USERNAME",
        nargs="?",
        default=0,
        help="Чат; без аргумента — вся БД.",
    )

    p = export_subparsers.add_parser(
        "xlsx",
        help="Экспорт в Excel.",
    )
    p.add_argument(
        "chat",
        type=chat_id_or_username,
        metavar="CHAT_ID|@USERNAME",
        nargs="?",
        default=0,
        help="ID чата для фильтрации; без аргумента — полный экспорт всей БД.",
    )

    return parser.parse_args()


@asynccontextmanager
async def async_ipc_lock(path: str | Path, timeout: float | None = 0) -> AsyncGenerator[None, None]:
    """Асинхронный контекстный менеджер для межпроцессной блокировки.
    Предотвращает одновременный запуск нескольких копий скрипта.

    path — путь к lock-файлу (например, my_script.lock)

    timeout:
      - 0 → не блокироваться (мгновенно вернуть результат)
      - None → ждать бесконечно
      - >0.0 → ждать указанное время в секундах
    """
    lock = fasteners.InterProcessLock(path)
    blocking = (timeout is None) or (timeout > 0)
    acquired = await asyncio.to_thread(lock.acquire, blocking=blocking, timeout=timeout)

    if not acquired:
        raise AlreadyRunningError(f"Скрипт уже запущен (lock-файл: {path})")
    log.debug(f"Lock-файл ({path}) успешно захвачен.")
    try:
        yield
    finally:
        await asyncio.to_thread(lock.release)
        log.debug(f"Lock-файл ({path}) освобождён.")


@contextmanager
def secure_umask(mask: int = 0o077) -> Generator[None, None, None]:
    """Контекстный менеджер для временной и безопасной установки umask процесса.
    Гарантирует восстановление исходной маски после выхода из блока.
    """
    original_umask = os.umask(mask)
    log.debug(f"Установлена временная umask={oct(mask)} для повышения безопасности.")
    try:
        yield
    finally:
        os.umask(original_umask)
        log.debug("Восстановлена исходная системная umask.")


def _format_bytes(size_bytes: int | float) -> str:
    """Форматирует байты в человекочитаемый вид (B, KiB, MiB, GiB, TiB)."""
    size_bytes = abs(size_bytes)
    for unit in ("B", "KiB", "MiB", "GiB"):
        if size_bytes < 1024.0:
            return f"{int(size_bytes)} {unit}" if unit == "B" else f"{size_bytes:.2f} {unit}"
        size_bytes /= 1024.0
    return f"{size_bytes:.2f} TiB"


def _format_duration(seconds: int | None) -> str:
    """Форматирует секунды в mm:ss (или h:mm:ss для длинных файлов)."""
    if seconds is None or seconds <= 0:
        return "00:00"

    seconds = int(seconds)

    m, s = divmod(seconds, 60)
    h, m = divmod(m, 60)

    if h > 0:
        return f"{h}:{m:02d}:{s:02d}"
    return f"{m:02d}:{s:02d}"


def _sanitize_filename(filename: str) -> str:
    """Очищает имя файла от запрещенных системных символов,
    сохраняя читаемость, пробелы и unicode (кириллицу, эмодзи).
    """
    MAX_FILENAME_BYTES = 215
    _RESERVED_NAMES = frozenset(
        {"CON", "PRN", "AUX", "NUL"} | {f"{p}{i}" for p in ("COM", "LPT") for i in range(1, 10)}
    )

    if not filename:
        return "unnamed_file"

    forbidden = '<>:"/\\|?*'
    remove = "".join(chr(c) for c in range(32)) + chr(127)
    table = str.maketrans(forbidden, "_" * len(forbidden), remove)

    cleaned = filename.translate(table).strip(" .")

    if not cleaned:
        return "unnamed_file"

    if cleaned.split(".")[0].upper() in _RESERVED_NAMES:
        cleaned = f"_{cleaned}"

    if len(cleaned.encode("utf-8")) > MAX_FILENAME_BYTES:
        stem, dot, ext = cleaned.rpartition(".")
        if dot and len(ext.encode("utf-8")) <= 20:
            ext_bytes = f".{ext}".encode()
            stem = stem.encode("utf-8")[: MAX_FILENAME_BYTES - len(ext_bytes)].decode(
                "utf-8", errors="ignore"
            )
            cleaned = f"{stem}.{ext}"
        else:
            cleaned = cleaned.encode("utf-8")[:MAX_FILENAME_BYTES].decode("utf-8", errors="ignore")

    return cleaned or "unnamed_file"


def _calculate_file_hash_sync(file_path: Path) -> str:
    """(СИНХРОННАЯ!) Вычисляет хэш-сумму BLAKE2b файла"""
    with open(file_path, "rb") as f:
        return hashlib.file_digest(f, "blake2b").hexdigest()


def _get_existing_parent(path: Path) -> Path:
    """Итеративно находит первый существующий родительский каталог."""
    current = path.resolve()
    while not current.exists() and current.parent != current:
        current = current.parent
    return current


def _get_size_safely(path: Path) -> int:
    """Возвращает размер обычного файла. Для симлинков/ошибок — 0."""
    try:
        st = path.lstat()
        if statmod.S_ISLNK(st.st_mode):
            log.debug(f"Игнорируется симлинк: {path}")
            return 0
        if statmod.S_ISREG(st.st_mode):
            return st.st_size
    except (FileNotFoundError, OSError) as e:
        if not isinstance(e, FileNotFoundError):
            log.debug(f"Не удалось получить размер файла '{path}': {e}")
    return 0


def remember_chat(chat: types.Chat) -> None:
    """Запоминает отображаемое имя чата."""
    name = chat.title or " ".join(
        p
        for p in (
            getattr(chat, "first_name", ""),
            getattr(chat, "last_name", ""),
        )
        if p
    )
    name = " ".join((name or "").split())
    CHAT_LABELS[chat.id] = (name, chat.username)


@functools.cache
def _username_from_session(chat_id: int) -> str | None:
    """Юзернейм из файла сессии (только чтение). None — не нашли/не смогли."""
    try:
        with sqlite3.connect(f"file:{SESSION_NAME}.session?mode=ro", uri=True) as conn:
            row = conn.execute(
                "SELECT username FROM usernames WHERE id = ? LIMIT 1", (chat_id,)
            ).fetchone()
            return row[0] if row and row[0] else None
    except sqlite3.Error:
        return None


@functools.cache
def _id_from_session(identifier: str) -> int | None:
    """Обратный резолв: @username / t.me-ссылка -> id из файла сессии."""
    raw = identifier.strip()

    if not raw:
        return None

    if raw.startswith("@"):
        uname = raw[1:]
    else:
        parsed = urlparse(raw if "://" in raw else f"https://{raw}")
        host = parsed.netloc.lower()

        if host in {"t.me", "telegram.me"}:
            parts = [p for p in parsed.path.split("/") if p]
            if not parts:
                return None
            if parts[0] == "s" and len(parts) > 1:
                parts = parts[1:]
            uname = parts[0]
        else:
            uname = raw

    uname = uname.strip().strip("/").lower()

    if not uname or uname.startswith("+") or uname in {"c", "joinchat"}:
        return None

    try:
        with sqlite3.connect(f"file:{SESSION_NAME}.session?mode=ro", uri=True) as conn:
            row = conn.execute(
                "SELECT id FROM usernames WHERE username = ? COLLATE NOCASE LIMIT 1",
                (uname,),
            ).fetchone()
            return row[0] if row else None
    except sqlite3.Error:
        return None


def chat_label(chat_id: int) -> str:
    """Человеческое имя чата для логов согласно chat_label_parts."""
    name, username = CHAT_LABELS.get(chat_id, ("", None))

    if username is None:
        username = _username_from_session(chat_id)

    values = {
        "title": name,
        "username": f"@{username}" if username else "",
        "id": str(chat_id),
    }

    parts = [values[key] for key in CHAT_LABEL_PARTS if values.get(key)]

    if not parts:
        return str(chat_id)

    if len(parts) == 1:
        return parts[0]

    has_title = "title" in CHAT_LABEL_PARTS and bool(name)
    if has_title:
        tail = [
            values[key]
            for key in CHAT_LABEL_PARTS
            if key != "title" and key in values and values[key]
        ]
        return f"{name} [{' | '.join(tail)}]" if tail else name

    return f"[{' | '.join(parts)}]"


def chat_id_or_username(identifier: str) -> ChatID:
    """Тип для argparse: числовой ID как есть, юзернейм — через файл сессии."""
    try:
        return int(identifier)
    except ValueError:
        pass
    if chat_id := _id_from_session(identifier):
        log.info(f"'{identifier}' найден в сессии: {chat_id}")
        return chat_id
    raise argparse.ArgumentTypeError(
        f"не удалось найти '{identifier}' в файле сессии. "
        f"Укажите числовой ID или запустите онлайн-режим (report/download), "
        f"чтобы чат попал в кэш сессии."
    )


# endregion

# region --- УПРАВЛЕНИЕ ДИСКОВЫМ ПРОСТРАНСТВОМ ---
# Блок, отвечающий исключительно за проверку наличия достаточного свободного места на диске. Включает обе стратегии проверки (статическую и динамическую).


async def check_disk_space() -> bool:
    """Главная функция-диспетчер для проверки свободного места на диске."""
    if MIN_FREE_SPACE_MB > 0:
        log.info("Выполняется СТАТИЧЕСКАЯ проверка свободного места...")
        return await asyncio.to_thread(
            _check_static_disk_space, Path(BACKUP_DIR), MIN_FREE_SPACE_MB
        )

    log.info("Выполняется ДИНАМИЧЕСКАЯ проверка свободного места...")
    return await _check_dynamic_disk_space()


def _check_static_disk_space(path_to_check: Path, required_mb: float) -> bool:
    """Проверяет наличие достаточного статического количества свободного места."""
    try:
        target_path = _get_existing_parent(path_to_check)
        if not target_path.exists():
            log.critical(
                f"Не удалось найти существующий путь для '{path_to_check}'. Проверьте, смонтирован ли диск."
            )
            return False

        _, _, free_bytes = shutil.disk_usage(target_path)
        log.info(
            f"Доступно на разделе '{target_path}': {_format_bytes(free_bytes)}. Требуется: {required_mb:.2f} МБ."
        )
        if free_bytes < (required_mb * 1024**2):
            log.critical("Недостаточно свободного места!")
            return False
        return True
    except Exception as e:
        log.critical(f"Не удалось проверить свободное место на диске. Ошибка: {e}")
        return False


async def _check_dynamic_disk_space() -> bool:
    """Выполняет умный динамический расчет необходимого места."""
    try:
        db_size, backups_size = await asyncio.to_thread(_scan_project_files_sync)

        if db_size == 0 and not Path(BACKUP_DIR).exists():
            log.info("Проект еще не содержит данных. Проверка свободного места не требуется.")
            return True

        safety_buffer_bytes = DYNAMIC_SPACE_SAFETY_BUFFER_MB * 1024 * 1024

        if BACKUP_ON_STARTUP:
            required_bytes = int(db_size * DYNAMIC_SPACE_COEFFICIENT + safety_buffer_bytes)
            log_reason = "Для создания бэкапа и роста БД"
        else:
            # Используем 20% от "запаса прочности" (coeff-1.0) для оценки роста.
            # Пример: coeff=1.5 (запас 50%) -> (1.5-1)*0.2+1 = 1.1 (запас 10%).
            growth_coefficient = (DYNAMIC_SPACE_COEFFICIENT - 1.0) * 0.2 + 1.0
            required_bytes = int(db_size * growth_coefficient + safety_buffer_bytes / 2)
            log_reason = "Для роста БД во время работы"

        backup_path = Path(BACKUP_DIR)
        target_path = _get_existing_parent(backup_path)
        _, _, free_bytes = await asyncio.to_thread(shutil.disk_usage, target_path)

        log.info(f"Текущий размер файлов проекта: {_format_bytes(db_size + backups_size)}.")
        log.info(f"  - Размер БД (с .wal/.shm): {_format_bytes(db_size)}.")
        log.info(f"  - Размер существующих бэкапов: {_format_bytes(backups_size)}.")
        log.info(f"{log_reason} требуется ~{_format_bytes(required_bytes)} свободного места.")
        log.info(f"Доступно на разделе '{target_path}': {_format_bytes(free_bytes)}.")

        if free_bytes < required_bytes:
            log.critical("Недостаточно свободного места!")
            return False

        return True
    except Exception as e:
        log.critical(
            f"Не удалось выполнить динамическую проверку места. Ошибка: {e}",
            exc_info=True,
        )
        return False


def _scan_project_files_sync() -> tuple[int, int]:
    """(СИНХРОННАЯ!) Безопасно сканирует файлы проекта, возвращая их размеры."""
    db_path = Path(DB_FILE)

    db_main_size = _get_size_safely(db_path)
    db_wal_size = _get_size_safely(db_path.with_name(f"{db_path.name}-wal"))
    db_shm_size = _get_size_safely(db_path.with_name(f"{db_path.name}-shm"))
    total_db_size = db_main_size + db_wal_size + db_shm_size

    total_backup_size = 0
    backup_path = Path(BACKUP_DIR)
    if backup_path.is_dir():
        for dirpath, _, filenames in os.walk(backup_path, followlinks=False):
            for filename in filenames:
                filepath = Path(dirpath) / filename
                total_backup_size += _get_size_safely(filepath)

    return total_db_size, total_backup_size


# endregion

# region --- УПРАВЛЕНИЕ РЕЗЕРВНЫМИ КОПИЯМИ ---
# Здесь собрана вся логика, связанная с созданием, ротацией, архивированием и удалением резервных копий базы данных.


async def create_database_backup() -> None:
    """Главная управляющая функция для процесса бэкапа."""
    source_db_path = Path(DB_FILE)
    backup_dir = Path(BACKUP_DIR)

    # --- БЛОК 1: Предварительные проверки (Хэш и существование) ---
    if not source_db_path.exists():
        log.debug(f"Файл БД '{source_db_path}' не существует, бэкап не требуется.")
        return

    current_db_hash = ""
    if BACKUP_ONLY_IF_CHANGED:
        log.info("Режим 'бэкап только при изменениях' активен (проверка по хэш-сумме).")
        hash_file_path = backup_dir / ".latest_backup.hash"

        try:
            log.info("Вычисляю хэш-сумму текущей БД (это может занять время для больших файлов)...")
            current_db_hash = await asyncio.to_thread(_calculate_file_hash_sync, source_db_path)
            log.debug(f"Текущий хэш БД: {current_db_hash[:40]}...")

            if not hash_file_path.is_file():
                log.info("Файл с хэшем предыдущего бэкапа не найден. Будет создан новый.")
            else:
                stored_hash = await asyncio.to_thread(hash_file_path.read_text, encoding="utf-8")
                log.debug(f"Хэш бэкапа БД: {stored_hash[:40]}...")
                if current_db_hash == stored_hash.strip():
                    log.info(
                        "Хэш-сумма БД не изменилась. Создание новой резервной копии пропущено."
                    )
                    return
                else:
                    log.info(
                        "Обнаружены изменения в БД (хэш-суммы не совпадают). Создание бэкапа необходимо."
                    )
        except Exception as e:
            log.warning(
                f"Не удалось проверить хэш-сумму БД. Ошибка: {e}. Бэкап будет создан для безопасности."
            )
            if not current_db_hash:
                current_db_hash = await asyncio.to_thread(_calculate_file_hash_sync, source_db_path)

    if not current_db_hash:
        current_db_hash = await asyncio.to_thread(_calculate_file_hash_sync, source_db_path)

    await asyncio.to_thread(os.makedirs, backup_dir, exist_ok=True)

    # --- БЛОК 2: Ротация "ДО" (если включена) ---
    if ROTATE_BEFORE_BACKUP:
        log.info("Режим экономии места: сначала ротация, потом создание бэкапа.")
        await _perform_rotation(source_db_path, backup_dir)

    # --- БЛОК 3: Создание бэкапа (Единая точка входа) ---
    # Мы вызываем создание здесь один раз, независимо от режима ротации
    new_backup_path = await _perform_backup_creation(source_db_path, backup_dir, current_db_hash)

    # --- БЛОК 4: Обработка результата и Ротация "ПОСЛЕ" ---
    if new_backup_path:
        # 4.1. Специфичная логика для max_backups=0
        # Если пользователь не хочет хранить "горячие" бэкапы, мы должны сразу же заархивировать созданный файл.
        if MAX_BACKUPS == 0 and ARCHIVE_OLD_BACKUPS:
            log.info(
                f"Настройка MAX_BACKUPS=0: Немедленная архивация свежего бэкапа '{new_backup_path.name}'..."
            )
            await _archive_backup_file(new_backup_path)

        # 4.2. Ротация "ПОСЛЕ", если она не была выполнена "ДО"
        if not ROTATE_BEFORE_BACKUP:
            log.info("Обычный режим: выполнение ротации после успешного создания бэкапа.")
            await _perform_rotation(source_db_path, backup_dir)
    else:
        # Если создание не удалось
        if not ROTATE_BEFORE_BACKUP:
            log.warning(
                "Создание бэкапа не удалось. Ротация старых копий пропущена для безопасности."
            )


async def _perform_backup_creation(
    source_db_path: Path, backup_dir: Path, db_hash: str
) -> Path | None:
    """Атомарно создает одну новую резервную копию.
    Возвращает путь к ней или None в случае ошибки.
    """
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    final_backup_path = backup_dir / f"{source_db_path.stem}_{timestamp}.sqlite.bak"
    tmp_backup_path = final_backup_path.with_suffix(final_backup_path.suffix + ".tmp")
    hash_file_path = backup_dir / ".latest_backup.hash"

    try:
        log.info("Создание копии БД...")
        async with aiosqlite.connect(source_db_path) as src_conn:
            async with src_conn.execute("PRAGMA integrity_check;") as cursor:
                result = await cursor.fetchone()
                if not result or result[0].lower() != "ok":
                    log.critical(
                        f"Исходная БД повреждена! integrity_check: '{result[0]}'. Бэкап отменен."
                    )
                    return None

            # Пишем во временный файл
            async with aiosqlite.connect(tmp_backup_path) as dest_conn:
                await src_conn.backup(dest_conn)

        # Если все прошло успешно, атомарно переименовываем .tmp в .bak
        await asyncio.to_thread(os.replace, tmp_backup_path, final_backup_path)

        # После успешного создания бэкапа, сохраняем хэш этого состояния
        await asyncio.to_thread(hash_file_path.write_text, db_hash, encoding="utf-8")
        log.debug(f"Сохранен новый хэш БД: {db_hash[:40]}...")

        log.info(f"Резервная копия '{final_backup_path.name}' успешно создана.")
        return final_backup_path

    except aiosqlite.Error as e:
        log.critical(f"Произошла ошибка базы данных при создании бэкапа: {e}")
        return None
    except OSError as e:
        log.critical(f"Произошла файловая ошибка при создании бэкапа: {e}")
        return None
    except Exception as e:
        log.critical(f"Непредвиденная ошибка при создании резервной копии БД: {e}.")
        return None

    finally:
        # В любом случае, если временный файл остался, удаляем его
        if await asyncio.to_thread(tmp_backup_path.exists):
            try:
                await asyncio.to_thread(os.remove, tmp_backup_path)
                log.warning(f"Удален временный файл бэкапа: {tmp_backup_path.name}")
            except OSError:
                pass


async def _perform_rotation(source_db_path: Path, backup_dir: Path) -> None:
    """Выполняет ротацию бэкапов и архивов согласно настройкам."""
    db_stem = source_db_path.stem

    # --- Ротация "горячих" бэкапов (.bak) ---
    # [ВАЖНО] Разрешаем вход даже если MAX_BACKUPS == 0, чтобы удалить "зависшие" файлы
    if MAX_BACKUPS >= 0:
        try:
            # Если лимит 0, то целевое количество файлов = 0.
            # Если лимит > 0, то вычисляем как обычно.
            if MAX_BACKUPS == 0:
                target_hot_backups = 0
            else:
                # Если ротация ДО создания, мы должны оставить место под 1 новый (MAX - 1).
                # Если ПОСЛЕ, то мы уже создали, значит храним ровно MAX.
                target_hot_backups = MAX_BACKUPS - 1 if ROTATE_BEFORE_BACKUP else MAX_BACKUPS

            # Защита от отрицательных чисел (на всякий случай)
            if target_hot_backups < 0:
                target_hot_backups = 0

            backups = await asyncio.to_thread(
                lambda: sorted(backup_dir.glob(f"{db_stem}_*.sqlite.bak"))
            )

            num_to_process = len(backups) - target_hot_backups

            if num_to_process > 0:
                log.info(
                    f"Ротация .bak: найдено {len(backups)}, лимит {target_hot_backups}. Обработка {num_to_process} лишних файлов..."
                )

                # Сортировка glob обычно по имени (а там дата), так что самые старые в начале списка.
                # Удаляем/Архивируем самые старые.
                for backup_path in backups[:num_to_process]:
                    if ARCHIVE_OLD_BACKUPS:
                        await _archive_backup_file(backup_path)
                    else:
                        log.debug(f"Удаляю старый бэкап: {backup_path.name}")
                        await asyncio.to_thread(os.remove, backup_path)
        except Exception as e:
            log.error(f"Ошибка при ротации бэкапов: {e}")

    # --- Ротация архивов (.xz) ---
    if MAX_ARCHIVES > 0 and ARCHIVE_OLD_BACKUPS:
        try:
            archives = await asyncio.to_thread(
                lambda: sorted(backup_dir.glob(f"{db_stem}_*.sqlite.bak.xz"))
            )
            num_to_delete = len(archives) - MAX_ARCHIVES
            if num_to_delete > 0:
                log.info(
                    f"Найдено {len(archives)} архивов (лимит: {MAX_ARCHIVES}). Удаляю {num_to_delete} самых старых..."
                )
                for archive_path in archives[:num_to_delete]:
                    await asyncio.to_thread(os.remove, archive_path)
                    log.debug(f"Удален старый архив: {archive_path.name}")
        except Exception as e:
            log.error(f"Ошибка при ротации архивов: {e}")


async def _archive_backup_file(backup_path: Path) -> None:
    """Атомарно сжимает один файл бэкапа и удаляет исходник."""
    archive_path = backup_path.with_suffix(backup_path.suffix + ".xz")
    tmp_path = archive_path.with_suffix(archive_path.suffix + ".tmp")
    try:
        log.info(f"Архивирую: {backup_path.name}...")
        await asyncio.to_thread(_compress_file_sync, backup_path, tmp_path, LZMA_PRESET)
        await asyncio.to_thread(os.replace, tmp_path, archive_path)
        await asyncio.to_thread(os.remove, backup_path)
        log.info(f" -> {archive_path.name}")
    except Exception as e:
        log.error(f"Не удалось заархивировать {backup_path.name}: {e}")
        if tmp_path.exists():
            await asyncio.to_thread(os.remove, tmp_path)


def _compress_file_sync(source_path: Path, dest_path: Path, preset: int) -> None:
    """(СИНХРОННАЯ!) Вспомогательная функция для сжатия файла с заданным пресетом."""
    with open(source_path, "rb") as f_in:
        with lzma.open(dest_path, "wb", preset=preset, check=lzma.CHECK_CRC64) as f_out:
            shutil.copyfileobj(f_in, f_out, length=1024 * 1024)  # type: ignore [arg-type]


# endregion

# region --- ОПЕРАЦИИ С БАЗОЙ ДАННЫХ ---
# Этот блок отвечает за все прямое взаимодействие с файлом SQLite: инициализация, подключение, валидация, ремонт и простые запросы (получение ID).


# todo Добавить версирование БД user_version
async def initialize_database() -> None:
    """Выполняется ОДИН РАЗ при запуске. Создает новую схему БД."""
    log.debug("Инициализация схемы базы данных...")
    async with aiosqlite.connect(DB_FILE) as conn:
        async with conn.execute("PRAGMA journal_mode = WAL;") as cursor:
            row = await cursor.fetchone()
        applied_mode = row[0] if row else None
        if not (applied_mode and applied_mode.lower() == "wal"):
            log.warning(
                f"Не удалось включить WAL, journal_mode: {applied_mode}. "
                "Возможно, БД на сетевой ФС."
            )
        # Основная таблица для хранения информации об аудиофайлах
        await conn.execute("""
            CREATE TABLE IF NOT EXISTS audios (
                chat_id INTEGER, message_id INTEGER, file_unique_id TEXT NOT NULL,
                file_name TEXT, file_size INTEGER, duration INTEGER,
                performer TEXT, title TEXT,
                PRIMARY KEY (chat_id, message_id)
            ) STRICT;
        """)
        # Индексы для быстрого поиска дубликатов
        await conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_chat_unique ON audios (chat_id, file_unique_id)"
        )
        await conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_chat_meta ON audios (chat_id, file_name, file_size, duration)"
        )

        # Таблица состояния синхронизации
        await conn.execute("""
            CREATE TABLE IF NOT EXISTS chat_sync_state (
                chat_id INTEGER PRIMARY KEY,
                is_fully_synced INTEGER NOT NULL DEFAULT 0,
                newest_scanned_id INTEGER DEFAULT 0
            ) STRICT;
        """)
        await conn.commit()
    log.debug("Схема базы данных готова.")


@asynccontextmanager
async def create_connection() -> AsyncGenerator[aiosqlite.Connection, None]:
    """Создает и возвращает НОВОЕ, полностью настроенное асинхронное соединение с БД."""
    log.debug("Создание нового оптимизированного соединения с БД...")

    async with aiosqlite.connect(DB_FILE) as conn:
        conn.row_factory = aiosqlite.Row

        pragmas = {
            "synchronous": "NORMAL",
            "temp_store": "MEMORY",
            "cache_size": DB_CACHE_SIZE,
            # рассмотреть: "mmap_size", "page_size", "wal_autocheckpoint" и т.д.
        }

        for pragma, value in pragmas.items():
            await conn.execute(f"PRAGMA {pragma} = {value};")

        log.debug(f"Настройки соединения успешно применены: {pragmas}")

        yield conn

    log.debug("Соединение с БД закрыто.")


async def validate_database() -> bool:
    """Проверяет целостность БД.
    Возвращает False (блокирует запуск), если:
    1. Файл БД физически поврежден.
    2. Отсутствуют обязательные таблицы.
    3. В таблице audios есть критически поврежденные данные (нужна команда repair).
    """
    log.info("Валидация базы данных...")
    warnings = 0

    try:
        async with aiosqlite.connect(DB_FILE) as conn:
            conn.row_factory = aiosqlite.Row

            # --- 1. Физическая целостность (CRITICAL) ---
            async with conn.execute("PRAGMA integrity_check;") as cur:
                result = await cur.fetchone()
                if not result or result[0].lower() != "ok":
                    log.critical(f"Файл БД физически поврежден: {result[0]}")
                    return False

            # --- 2. Наличие таблиц (CRITICAL) ---
            async with conn.execute("SELECT name FROM sqlite_master WHERE type='table'") as cur:
                tables = {row[0] for row in await cur.fetchall()}

            required = {"audios", "chat_sync_state"}
            missing = required - tables
            if missing:
                log.critical(f"Отсутствуют обязательные таблицы: {missing}")
                return False

            # --- 3. Целостность данных в audios (CRITICAL) ---
            # file_name IS NULL сам по себе валиден (у аудио есть performer/title);
            # сломанной считаем запись без всех трёх текстовых полей сразу
            async with conn.execute("""
                SELECT COUNT(*) FROM audios 
                WHERE file_unique_id IS NULL OR file_unique_id = '' 
                   OR (file_name IS NULL AND performer IS NULL AND title IS NULL)
                   OR file_size IS NULL OR file_size < 0 
                   OR duration IS NULL OR duration < 0
            """) as cur:
                broken_count = (await cur.fetchone())[0]
                if broken_count > 0:
                    log.critical(
                        f"ОБНАРУЖЕНО {broken_count} ПОВРЕЖДЁННЫХ ЗАПИСЕЙ в таблице 'audios'.\n"
                        "Запуск поиска дубликатов на таких данных опасен (возможна потеря данных).\n"
                        "Пожалуйста, запустите скрипт с командой: repair"
                    )
                    return False

            # --- 4. Наличие индексов (WARNING) ---
            async with conn.execute("SELECT name FROM sqlite_master WHERE type='index'") as cur:
                indexes = {row[0] for row in await cur.fetchall()}

            for idx in ("idx_chat_unique", "idx_chat_meta"):
                if idx not in indexes:
                    warnings += 1
                    log.warning(f"Отсутствует индекс '{idx}'. Это замедлит работу.")

            # --- 5. Второстепенные проблемы (WARNING) ---
            async with conn.execute("""
                        SELECT COUNT(*) FROM chat_sync_state WHERE newest_scanned_id < 0
                    """) as cur:
                bad_cursors = (await cur.fetchone())[0]

                if bad_cursors > 0:
                    warnings += 1
                    log.warning(
                        f"Найдено {bad_cursors} некорректных курсоров истории (рекомендуется команда repair)."
                    )

    except aiosqlite.Error as e:
        log.critical(f"Ошибка SQLite при валидации: {e}")
        return False
    except Exception as e:
        log.critical(f"Непредвиденная ошибка валидации: {e}")
        return False

    if warnings:
        log.warning(f"Валидация прошла с {warnings} предупреждениями.")
    else:
        log.info("Валидация базы данных успешно пройдена.")

    return True


async def repair_database(app: Client) -> None:
    """Выполняет умное восстановление и очистку базы данных.
    Пытается восстановить поврежденные записи, используя данные из Telegram.
    """
    log.info("=" * 15 + "ЗАПУСК РЕМОНТА БД" + "=" * 15)

    # --- ЧАСТЬ 1: Работа с данными (Исправление и Очистка) ---
    async with aiosqlite.connect(DB_FILE) as conn:
        conn.row_factory = aiosqlite.Row

        async with conn.execute("PRAGMA integrity_check;") as cur:
            result = await cur.fetchone()
            if not result or result[0].lower() != "ok":
                log.critical("БД физически повреждена. Восстановите из бэкапа вручную.")
                return

        log.info("Этап 1: Поиск и попытка восстановления поврежденных записей...")

        # file_name IS NULL сам по себе валиден; ищем записи без всех трёх текстовых полей
        query = """
            SELECT chat_id, message_id FROM audios 
            WHERE file_unique_id IS NULL OR file_unique_id = '' 
               OR (file_name IS NULL AND performer IS NULL AND title IS NULL)
               OR file_size IS NULL OR file_size < 0 
               OR duration IS NULL OR duration < 0
        """
        async with conn.execute(query) as cursor:
            broken_records = await cursor.fetchall()

        if not broken_records:
            log.info("Поврежденных записей для восстановления не найдено.")
        else:
            log.info(
                f"Найдено {len(broken_records)} потенциально поврежденных записей. Начинаю сверку с Telegram..."
            )

            records_by_chat = defaultdict(list)
            for rec in broken_records:
                records_by_chat[rec["chat_id"]].append(rec["message_id"])

            records_to_update = []
            ids_to_delete = []
            semaphore = asyncio.Semaphore(VERIFY_CONCURRENCY)

            async def fetch_and_process_chunk(chat_id, chunk_ids):
                async with semaphore:
                    try:
                        messages = await app.get_messages(chat_id, chunk_ids)
                        for original_id, msg in zip(chunk_ids, messages, strict=True):
                            attrs = _get_audio_attributes(msg)
                            if attrs:
                                records_to_update.append((*attrs, chat_id, original_id))
                            else:
                                ids_to_delete.append((chat_id, original_id))
                    except Exception as e:
                        log.error(
                            f"Не удалось проверить чанк для чата {chat_label(chat_id)} (ID: {chunk_ids[0]}...): {e}. Чанк будет пропущен."
                        )

            tasks = []
            for chat_id, msg_ids in records_by_chat.items():
                for chunk_tuple in itertools.batched(msg_ids, VERIFY_CHUNK_SIZE):
                    tasks.append(fetch_and_process_chunk(chat_id, list(chunk_tuple)))

            if tasks:
                await asyncio.gather(*tasks)

            if records_to_update:
                await conn.executemany(
                    "UPDATE audios SET file_unique_id=?, file_name=?, file_size=?, duration=?, performer=?, title=? WHERE chat_id=? AND message_id=?",
                    records_to_update,
                )
                log.info(f"  Восстановлено (обновлено) {len(records_to_update)} записей.")

            if ids_to_delete:
                await conn.executemany(
                    "DELETE FROM audios WHERE chat_id = ? AND message_id = ?",
                    ids_to_delete,
                )
                log.info(f"  Удалено {len(ids_to_delete)} невосстановимых записей.")

        log.info("\nЭтап 2: Выполнение стандартной очистки...")
        async with conn.execute(
            "UPDATE chat_sync_state SET newest_scanned_id = 0 WHERE newest_scanned_id < 0"
        ) as cursor:
            if cursor.rowcount > 0:
                log.info(f"  Сброшено {cursor.rowcount} некорректных newest_scanned_id.")

        log.info("  Проверка и создание индексов...")
        await conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_chat_unique ON audios (chat_id, file_unique_id)"
        )
        await conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_chat_meta ON audios (chat_id, file_name, file_size, duration)"
        )
        log.info("  Индексы проверены.")

        await conn.commit()

    await asyncio.sleep(0.1)

    # --- ЧАСТЬ 2: Оптимизация (VACUUM) на новом соединении ---
    log.info("\nЭтап 3: Оптимизация файла БД (VACUUM)...")
    try:
        async with aiosqlite.connect(DB_FILE, isolation_level=None) as conn:
            log.info("  -> Выполнение wal_checkpoint(TRUNCATE)...")
            async with conn.execute("PRAGMA wal_checkpoint(TRUNCATE);") as cursor:
                await cursor.fetchall()
            log.info("  -> Checkpoint успешно выполнен, курсор закрыт.")

            log.info("  -> Выполнение VACUUM...")
            await conn.execute("VACUUM;")

            log.info("  Оптимизация и сжатие файла БД завершены.")

    except aiosqlite.Error as e:
        log.error(f"  Произошла ошибка на этапе оптимизации БД: {e}")

    log.info("=" * 15 + "РЕМОНТ БД ЗАВЕРШЕН" + "=" * 15)


# endregion

# region --- Под-блок: Экспорты ---
# todo Поиск по всем чатам в БД с пониманием опечаток(--search)


# todo ротация?
def _build_export_path(
    chat_id: ChatID,
    kind: str,
    ext: str,
    ts: str | None = None,
) -> Path:
    """Строит путь файла экспорта: exports/<chat>/<ts>_<kind>.<ext>.

    chat_id == 0 — полный экспорт всей БД, кладётся в exports/_full с именем БД
    в имени файла. chat_id хранится "сырым". ts по умолчанию — текущий момент
    (секунды); для прогонов по нескольким чатам считается один раз и пробрасывается.
    """
    if ts is None:
        ts = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")

    exports_root = Path(EXPORTS_DIR)
    if chat_id == 0:
        sub = exports_root / "_full"
        filename = f"{ts}_{kind}_of_{Path(DB_FILE).stem}.{ext}"
    else:
        sub = exports_root / str(chat_id)
        filename = f"{ts}_{kind}.{ext}"

    sub.mkdir(parents=True, exist_ok=True)
    return sub / filename


async def _generic_export_to_txt(
    chat_id: ChatID, output_file: str, sql_query: str, line_formatter: RowFormatter
) -> None:
    """Общая (generic) функция для экспорта данных из БД в текстовый файл.
    Принимает SQL-запрос и функцию-форматтер для строк.
    """
    log.info(f"Запущена задача экспорта для чата {chat_label(chat_id)} в файл '{output_file}'...")

    try:
        if not Path(DB_FILE).exists():
            log.critical(f"Файл базы данных '{DB_FILE}' не найден. Нечего экспортировать.")
            return

        async with aiosqlite.connect(DB_FILE) as conn:
            conn.row_factory = aiosqlite.Row
            async with conn.execute(sql_query, (chat_id,)) as cursor:
                data_rows = await cursor.fetchall()

        if not data_rows:
            log.warning(f"В базе данных не найдено записей для чата {chat_label(chat_id)}.")
            return

        log.info(f"Найдено {len(data_rows)} записей. Записываю в файл...")

        def write_to_file_sync():
            """Синхронная функция для записи в файл."""
            with open(output_file, "w", encoding="utf-8") as f:
                for row in data_rows:
                    line = line_formatter(row)
                    if line is not None:
                        f.write(line + "\n")

        await asyncio.to_thread(write_to_file_sync)
        log.info(f"Экспорт успешно завершен. Файл сохранен как '{output_file}'.")

    except aiosqlite.Error as e:
        log.critical(f"Произошла ошибка SQLite при экспорте: {e}")
    except Exception as e:
        log.critical(f"Произошла непредвиденная ошибка при экспорте: {e}", exc_info=True)


async def _generic_export_to_csv(
    chat_id: ChatID,
    kind: str,
    sql_query: str,
    sql_query_full: str,
    header: list[str],
    row_formatter: CsvRowFormatter,
) -> None:
    """Общая функция для экспорта данных из БД в CSV.

    chat_id == 0 — полный экспорт всей БД (используется sql_query_full
    без параметров). Форматтер возвращает список полей строки или None,
    чтобы пропустить запись.
    """
    is_full_export = chat_id == 0
    output_file = _build_export_path(chat_id, kind, "csv")

    if is_full_export:
        log.info(f"Запущена задача ПОЛНОГО экспорта '{kind}' в '{output_file}'...")
    else:
        log.info(
            f"Запущена задача экспорта '{kind}' для чата {chat_label(chat_id)} в '{output_file}'..."
        )

    try:
        if not Path(DB_FILE).exists():
            log.critical(f"Файл базы данных '{DB_FILE}' не найден. Нечего экспортировать.")
            return

        async with aiosqlite.connect(DB_FILE) as conn:
            conn.row_factory = aiosqlite.Row
            if is_full_export:
                query, params = sql_query_full, ()
            else:
                query, params = sql_query, (chat_id,)
            async with conn.execute(query, params) as cursor:
                data_rows = await cursor.fetchall()

        if not data_rows:
            target = "базе данных" if is_full_export else f"чате {chat_label(chat_id)}"
            log.warning(f"В {target} не найдено записей.")
            return

        log.info(f"Найдено {len(data_rows)} записей. Обработка и сохранение в CSV...")

        def write_to_file_sync():
            import csv

            # utf-8-sig нужен, чтобы Excel автоматически правильно распознал кириллицу
            with open(output_file, "w", encoding="utf-8-sig", newline="") as f:
                # delimiter=';' используется по умолчанию в русскоязычном Excel
                writer = csv.writer(f, delimiter=";")
                writer.writerow(header)
                for row in data_rows:
                    fields = row_formatter(row)
                    if fields is not None:
                        writer.writerow(fields)

        await asyncio.to_thread(write_to_file_sync)
        log.info(f"Экспорт успешно завершен. Файл сохранен как '{output_file}'.")

    except aiosqlite.Error as e:
        log.critical(f"Произошла ошибка SQLite при экспорте: {e}")
    except Exception as e:
        log.critical(f"Произошла непредвиденная ошибка при экспорте: {e}", exc_info=True)


async def export_filenames_to_txt(chat_id: ChatID) -> None:
    """Экспортирует ТОЛЬКО имена файлов. (Функция-обертка)"""
    output_file = _build_export_path(chat_id, "filenames", "txt")

    await _generic_export_to_txt(
        chat_id=chat_id,
        output_file=str(output_file),
        sql_query="SELECT file_name FROM audios WHERE chat_id = ? ORDER BY file_name",
        line_formatter=lambda row: row["file_name"] if row["file_name"] else None,
    )


async def export_filenames_with_url_to_txt(chat_id: ChatID) -> None:
    """Экспортирует имена файлов и ССЫЛКИ. (Функция-обертка)"""
    try:
        from wcwidth import wcswidth
    except ImportError:
        log.warning("Библиотека 'wcwidth' не найдена. Выравнивание колонок может быть неточным.")
        wcswidth = len

    # if "-100" not in str(chat_id): log.warning("Возможно личный чат, ссылки могут быть не действительны!")
    if chat_id >= 0:
        log.warning("Возможно личный чат, ссылки могут быть не действительны!")
    public_chat_id = str(chat_id).removeprefix("-100")

    def formatter(row: aiosqlite.Row) -> str | None:
        if not row["file_name"]:
            return None
        file_name = row["file_name"]
        message_id = row["message_id"]
        target_width = 80  # Целевая визуальная ширина колонки

        # 1. Вычисляем реальную визуальную ширину имени файла
        visual_width = wcswidth(file_name)
        if visual_width < 0:
            visual_width = len(file_name)

        # 2. Вычисляем, сколько пробелов нужно добавить
        padding_needed = target_width - visual_width

        # 3. Если имя файла уже длиннее нашей колонки, добавим всего один пробел
        if padding_needed <= 0:
            padding_needed = 1

        padding = " " * padding_needed

        # 4. Собираем строку
        return f"{file_name}{padding}| https://t.me/c/{public_chat_id}/{message_id}"

    output_file = _build_export_path(chat_id, "filenames_with_urls", "txt")

    await _generic_export_to_txt(
        chat_id=chat_id,
        output_file=str(output_file),
        sql_query="SELECT file_name, message_id FROM audios WHERE chat_id = ? ORDER BY file_name",
        line_formatter=formatter,
    )


async def export_cleaned_names_to_csv(chat_id: ChatID) -> None:
    """Экспортирует процесс очистки имен файлов в CSV для проверки работы фильтров.
    Формат: Исходное имя, После _clean_filename, После default_process
    Если chat_id == 0, экспортируется вся база целиком.
    """

    def formatter(row: DBRow) -> list[str] | None:
        orig = row["file_name"]
        if not orig:
            return None
        cleaned = _clean_filename(orig)
        return [orig, cleaned, _process_for_fuzzy(cleaned)]

    await _generic_export_to_csv(
        chat_id=chat_id,
        kind="cleaned_names",
        sql_query="SELECT file_name FROM audios WHERE chat_id = ? AND file_name IS NOT NULL ORDER BY file_name",
        sql_query_full="SELECT file_name FROM audios WHERE file_name IS NOT NULL ORDER BY file_name",
        header=["Исходное имя", "После _clean_filename", "После default_process"],
        row_formatter=formatter,
    )


async def export_cleaned_meta_to_csv(chat_id: ChatID) -> None:
    """Экспортирует процесс очистки метаданных (performer+title) в CSV.
    Формат: Performer, Title, После _clean_meta, После default_process
    Показывает ровно ту строку, которую видит fuzzy-матчер.
    Если chat_id == 0, экспортируется вся база целиком.
    """

    def formatter(row: DBRow) -> list[str] | None:
        cleaned = _clean_meta(row["performer"], row["title"])
        if not cleaned:
            return None
        return [
            row["performer"] or "",
            row["title"] or "",
            cleaned,
            _process_for_fuzzy(cleaned),
        ]

    where = "(performer IS NOT NULL OR title IS NOT NULL)"
    order = "ORDER BY performer, title"
    await _generic_export_to_csv(
        chat_id=chat_id,
        kind="cleaned_meta",
        sql_query=f"SELECT performer, title FROM audios WHERE chat_id = ? AND {where} {order}",
        sql_query_full=f"SELECT performer, title FROM audios WHERE {where} {order}",
        header=["Performer", "Title", "После _clean_meta", "После default_process"],
        row_formatter=formatter,
    )


async def download_chat_audio(app: Client, chat_id: ChatID) -> None:
    """Скачивает все аудиофайлы из указанного чата в локальную папку downloads."""
    log.info(f"Запуск режима СКАЧИВАНИЯ для чата {chat_label(chat_id)}...")

    download_dir = Path(DOWNLOADS_DIR) / str(chat_id)
    download_dir.mkdir(parents=True, exist_ok=True)
    log.info(f"Папка для сохранения: {download_dir.resolve()}")

    IS_TTY = sys.stdout.isatty()

    async with (
        create_connection() as conn,
        conn.execute(
            "SELECT message_id, file_name, file_size FROM audios WHERE chat_id = ? ORDER BY message_id",
            (chat_id,),
        ) as cursor,
    ):
        records = await cursor.fetchall()

    if not records:
        log.warning(
            f"В базе данных нет записей для чата {chat_label(chat_id)}. Сначала запустите синхронизацию."
        )
        return

    total_files = len(records)

    total_expected_bytes = sum((r["file_size"] or 0) for r in records)

    _, _, free_bytes = await asyncio.to_thread(shutil.disk_usage, download_dir)

    log.info(f"Статистика для скачивания (Чат {chat_label(chat_id)}):")
    log.info(f"  - Файлов в БД: {total_files}")
    log.info(f"  - Общий размер: {_format_bytes(total_expected_bytes)} (без учета уже скачанных)")
    log.info(f"  - Свободно на диске: {_format_bytes(free_bytes)}")

    if free_bytes < total_expected_bytes:
        log.warning("Свободного места на диске меньше, чем суммарный размер файлов!")
        log.warning("Если часть файлов уже была скачана ранее, они будут пропущены.")
        log.warning("Загрузка начнется через 10 секунд... Нажмите Ctrl+C для отмены.")

        try:
            await asyncio.sleep(10)
        except asyncio.CancelledError:
            log.info("Загрузка отменена.")
            raise

    log.info("Инициализация очереди загрузки...")

    # Ограничение одновременных загрузок
    # todo (Проверить на премиуме)
    DOWNLOAD_CONCURRENCY = 6 if app.me.is_premium else 3
    queue = asyncio.Queue(maxsize=16)

    stats = {"success": 0, "skipped": 0, "error": 0}
    active_downloads: dict[str, int] = {}

    def _clear_line():
        if IS_TTY:
            sys.stdout.write("\r\033[K")
            sys.stdout.flush()

    def display_status():
        if not IS_TTY:
            return

        parts = []
        for name in sorted(active_downloads.keys()):
            percent = active_downloads[name]
            short_name = (name[:20] + "..") if len(name) > 23 else name
            parts.append(f"[{short_name}: {percent}%]")

        status_line = "  ".join(parts)

        sys.stdout.write(
            f"\r\033[KЗагрузка ({len(active_downloads)}/{DOWNLOAD_CONCURRENCY} парал.): {status_line}"
        )
        sys.stdout.flush()

    def create_progress_callback(filename: str):
        last_update_time = 0

        def progress(current, total):
            nonlocal last_update_time
            if total == 0:
                return

            percent = int(current * 100 / total)
            now = time.time()

            if IS_TTY and (percent >= 100 or (now - last_update_time > 0.1)):
                active_downloads[filename] = percent
                display_status()
                last_update_time = now

        return progress

    async def worker():
        while True:
            try:
                task_item = await queue.get()
            except asyncio.CancelledError:
                return

            message, file_name, expected_size = task_item
            safe_name = "unknown"
            final_path = None

            try:
                # --- 1. Подготовка имени файла ---
                base_name = file_name if file_name else f"audio_{message.id}"
                safe_name = _sanitize_filename(base_name)

                path_obj = Path(safe_name)
                # Если расширения нет, пытаемся угадать по mime-type
                if not path_obj.suffix:
                    mime = None
                    if message.audio:
                        mime = message.audio.mime_type
                    elif message.document:
                        mime = message.document.mime_type
                    guessed_ext = app.guess_extension(mime) if mime else None
                    safe_name += guessed_ext if guessed_ext else ".mp3"

                final_path = download_dir / safe_name

                if safe_name in active_downloads:
                    safe_name = f"{message.id}_{safe_name}"
                    final_path = download_dir / safe_name

                should_download = True
                if final_path.exists():
                    existing_size = _get_size_safely(final_path)

                    try:
                        existing_mtime = final_path.stat().st_mtime
                    except OSError:
                        existing_mtime = 0

                    expected_mtime = message.date.timestamp() if message.date else 0

                    size_matches = existing_size > 0 and abs(existing_size - expected_size) < 100
                    # 2 сек погрешности для файловых систем типа FAT32/exFAT
                    mtime_matches = (
                        expected_mtime == 0 or abs(existing_mtime - expected_mtime) <= 2.0
                    )

                    if size_matches and mtime_matches:
                        should_download = False
                    else:
                        safe_name = f"{message.id}_{safe_name}"
                        final_path = download_dir / safe_name

                        if final_path.exists():
                            existing_size_renamed = _get_size_safely(final_path)
                            try:
                                existing_mtime_renamed = final_path.stat().st_mtime
                            except OSError:
                                existing_mtime_renamed = 0

                            if (
                                existing_size_renamed > 0
                                and abs(existing_size_renamed - expected_size) < 100
                                and (
                                    expected_mtime == 0
                                    or abs(existing_mtime_renamed - expected_mtime) <= 2.0
                                )
                            ):
                                should_download = False

                # --- 3. Выполнение действия ---
                if not should_download:
                    if not IS_TTY:
                        log.debug(f"Файл существует, пропуск: {safe_name}")
                    stats["skipped"] += 1
                else:
                    active_downloads[safe_name] = 0
                    # --- Скачивание ---
                    if not IS_TTY:
                        log.info(f"Начало загрузки: {safe_name} ({_format_bytes(expected_size)})")
                    else:
                        display_status()

                    progress_callback = create_progress_callback(safe_name)

                    await app.download_media(
                        message,
                        file_name=str(final_path),
                        progress=progress_callback,
                    )

                    # --- Установка оригинальной даты изменения файла ---
                    if message.date:
                        try:
                            mtime = message.date.timestamp()
                            os.utime(final_path, (mtime, mtime))
                        except Exception as e:
                            log.debug(f"Не удалось обновить дату для файла {safe_name}: {e}")

                    if not IS_TTY:
                        log.info(f"Успешно скачано: {safe_name}")

                    stats["success"] += 1

            except Exception as e:
                if final_path and final_path.exists():
                    try:
                        os.remove(final_path)
                    except OSError:
                        pass

                _clear_line()
                log.error(f"Ошибка загрузки {safe_name}: {e}")
                stats["error"] += 1

            finally:
                active_downloads.pop(safe_name, None)
                display_status()
                queue.task_done()

    # Запуск воркеров
    workers = [asyncio.create_task(worker()) for _ in range(DOWNLOAD_CONCURRENCY)]

    # Producer
    chunk_size = 100
    local_meta = {r["message_id"]: (r["file_name"], r["file_size"]) for r in records}
    all_msg_ids = list(local_meta.keys())
    processed_count = 0

    try:
        for chunk_ids in itertools.batched(all_msg_ids, chunk_size):
            try:
                messages = await app.get_messages(chat_id, list(chunk_ids))
            except Exception as e:
                _clear_line()
                log.error(f"Неустранимая ошибка при получении списка сообщений: {e}")
                continue

            if not messages:
                continue

            for msg in messages:
                if not msg or msg.empty:
                    continue
                if not (msg.audio or msg.document):
                    continue

                db_name, db_size = local_meta.get(msg.id, (None, 0))

                current_file_name = None
                if msg.audio and msg.audio.file_name:
                    current_file_name = msg.audio.file_name
                elif msg.document and msg.document.file_name:
                    current_file_name = msg.document.file_name

                if not current_file_name:
                    current_file_name = db_name

                await queue.put((msg, current_file_name, db_size))

            processed_count += len(chunk_ids)
            if processed_count % 500 == 0:
                _clear_line()
                log.info(f"--- Обработано метаданных {processed_count}/{total_files} ---")
                display_status()

        await queue.join()

    finally:
        for w in workers:
            w.cancel()
        await asyncio.gather(*workers, return_exceptions=True)

    _clear_line()
    log.info(
        f"\n{'=' * 20}\nСкачивание завершено.\n"
        f"Успешно: {stats['success']}\n"
        f"Пропущено: {stats['skipped']}\n"
        f"Ошибок: {stats['error']}"
    )


async def export_database_to_xlsx(chat_id: ChatID) -> None:
    """Универсальный экспорт БД в Excel.
    Если chat_id == 0, экспортируется вся база целиком.
    Иначе — данные фильтруются по конкретному чату.
    """
    try:
        import openpyxl
        from openpyxl.utils.cell import get_column_letter
    except ImportError:
        log.critical(
            "Для экспорта в Excel требуется библиотека openpyxl.\n"
            "Пожалуйста, установите её: pip install openpyxl"
        )
        return

    # Константы
    EXCEL_MAX_SHEET_NAME_LEN: Final[int] = 31
    EXCEL_MAX_COLUMN_WIDTH: Final[int] = 80
    EXCEL_PADDING: Final[int] = 2
    INVALID_EXCEL_CHARS: Final[frozenset[str]] = frozenset("[]:*?/\\")

    # Определяем имя файла и режим
    is_full_export = chat_id == 0
    output_file = _build_export_path(chat_id, "database_export", "xlsx")

    if is_full_export:
        log.info(f"Запущена задача ПОЛНОГО экспорта базы данных в '{output_file}'...")
    else:
        log.info(f"Запущена задача экспорта для чата {chat_id} в '{output_file}'...")

    if not Path(DB_FILE).exists():
        log.critical(f"Файл базы данных '{DB_FILE}' не найден.")
        return

    # --- ВНУТРЕННИЕ ХЕЛПЕРЫ ---

    def _sanitize_sheet_name(name: str) -> str:
        return "".join("_" if c in INVALID_EXCEL_CHARS else c for c in name)

    def _get_unique_sheet_name(base_name: str, used_names: set[str]) -> str:
        clean_name = _sanitize_sheet_name(base_name)
        safe_base = clean_name[:EXCEL_MAX_SHEET_NAME_LEN]

        if safe_base not in used_names:
            used_names.add(safe_base)
            return safe_base

        for i in range(1, 1000):
            suffix = f"_{i}"
            allowed_len = EXCEL_MAX_SHEET_NAME_LEN - len(suffix)
            candidate = clean_name[:allowed_len] + suffix
            if candidate not in used_names:
                used_names.add(candidate)
                return candidate
        return safe_base[:EXCEL_MAX_SHEET_NAME_LEN]

    def _convert_cell_value(value: Any) -> Any:
        if value is None:
            return ""
        if isinstance(value, bytes):
            return f"<BLOB {len(value)} bytes>"
        if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
            return f"'{value}"
        return value

    def _append_table_to_workbook(
        wb: openpyxl.Workbook,
        table_name: str,
        headers: list[str],
        rows: list[tuple],
        used_sheet_names: set[str],
    ):
        sheet_title = _get_unique_sheet_name(table_name, used_sheet_names)
        ws = wb.create_sheet(title=sheet_title)
        ws.append(headers)

        for row in rows:
            safe_row = [_convert_cell_value(cell) for cell in row]
            ws.append(safe_row)

        for i, col_cells in enumerate(ws.columns, start=1):
            max_len = 0
            column_letter = get_column_letter(i)
            for cell in col_cells[:50]:
                try:
                    val = cell.value
                    if val is None:
                        val_len = 0
                    elif isinstance(val, (datetime.datetime, datetime.date)):
                        val_len = 18
                    else:
                        val_len = len(str(val))
                    if val_len > max_len:
                        max_len = val_len
                except Exception:
                    continue

            adjusted_width = min(max_len + EXCEL_PADDING, EXCEL_MAX_COLUMN_WIDTH)
            ws.column_dimensions[column_letter].width = adjusted_width

    # --- ОСНОВНАЯ ЛОГИКА ---

    wb = openpyxl.Workbook()
    if wb.active:
        wb.remove(wb.active)

    used_sheet_names: set[str] = set()
    data_found = False

    try:
        async with aiosqlite.connect(DB_FILE) as conn:
            async with conn.execute("SELECT name FROM sqlite_master WHERE type='table';") as cursor:
                tables = await cursor.fetchall()
                table_names = [row[0] for row in tables]

            log.info(f"Найдено таблиц: {len(table_names)}.")

            for table in table_names:
                if table.startswith("sqlite_"):
                    continue

                try:
                    safe_table_name = table.replace('"', '""')
                    async with conn.execute(f'PRAGMA table_info("{safe_table_name}");') as cursor:
                        columns_info = await cursor.fetchall()
                        column_names = [col[1] for col in columns_info]
                except Exception as e:
                    log.error(f"Не удалось получить схему таблицы '{table}': {e}")
                    continue

                has_chat_id = "chat_id" in column_names

                if not is_full_export and has_chat_id:
                    query = f'SELECT * FROM "{safe_table_name}" WHERE chat_id = ?'
                    params = (chat_id,)
                else:
                    query = f'SELECT * FROM "{safe_table_name}"'
                    params = ()

                try:
                    async with conn.execute(query, params) as cursor:
                        db_rows = await cursor.fetchall()
                        rows_tuples = [tuple(row) for row in db_rows]

                        if not rows_tuples:
                            if is_full_export:
                                log.info(f"  -> Таблица '{table}': пуста.")
                            else:
                                log.debug(f"  -> Таблица '{table}': нет данных для этого чата.")
                            continue

                        data_found = True
                        log.info(f"  -> Таблица '{table}': экспорт {len(rows_tuples)} строк...")

                        await asyncio.to_thread(
                            _append_table_to_workbook,
                            wb,
                            table,
                            column_names,
                            rows_tuples,
                            used_sheet_names,
                        )

                except Exception as e:
                    log.error(f"Ошибка при обработке таблицы '{table}': {e}")

        if not data_found:
            log.warning("Данные для экспорта не найдены.")
            return

        log.info("Сохранение файла на диск...")
        await asyncio.to_thread(wb.save, output_file)
        log.info(f"Экспорт успешно завершен. Файл: '{output_file}'")

    except Exception as e:
        log.critical(f"Критическая ошибка при экспорте Excel: {e}", exc_info=True)
        if Path(output_file).exists():
            try:
                os.remove(output_file)
            except OSError:
                pass


async def create_duplicates_report(
    chat_id: ChatID, conn: aiosqlite.Connection, ts: str | None = None
) -> None:
    """Создает человекочитаемый отчет о дубликатах с ссылками."""
    log.info(f"Генерация отчета по дубликатам для чата {chat_label(chat_id)}...")

    groups, edge_meta = await _get_potential_duplicate_groups(chat_id, conn)

    if not groups:
        log.info("Дубликатов не найдено. Отчет не нужен.")
        return

    # Раскладываем рёбра по группам один раз: message_id -> индекс группы.
    id_to_group_idx: dict[int, int] = {}
    for gi, group in enumerate(groups):
        for row in group:
            id_to_group_idx[row["message_id"]] = gi

    edges_by_group: defaultdict[int, list[tuple[int, int, EdgeInfo]]] = defaultdict(list)
    for (a, b), info in edge_meta.items():
        gi = id_to_group_idx.get(a)
        if gi is not None and gi == id_to_group_idx.get(b):
            edges_by_group[gi].append((a, b, info))

    if chat_id >= 0:
        log.warning("Возможно личный чат, ссылки могут быть не действительны!")
    clean_chat_id = str(chat_id).removeprefix("-100")

    report_file = _build_export_path(chat_id, "report_duplicates", "txt", ts=ts)

    buf = io.StringIO()

    buf.write(f"ОТЧЕТ О ДУБЛИКАТАХ (Чат: {chat_label(chat_id)})\n")
    buf.write(f"Дата генерации: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
    buf.write(f"Найдено групп: {len(groups)}\n\n")
    pretty_priority = ", ".join(f"{n} ~{t:.0%}" if t else n for n, t in KEEP_PRIORITY)
    buf.write(f"Стратегия оригинала: {pretty_priority}\n")
    buf.write(
        "Пометка [KEEP] предварительная: если кандидат не пройдёт верификацию\n"
        "(удалён/изменён в Telegram), оригиналом станет следующий в группе.\n"
    )

    if ENABLE_FUZZY_MATCHING:
        buf.write("\n[НАСТРОЙКИ FUZZY ПОИСКА]\n")
        buf.write(f"  • Режим:          {FUZZY_MATCHING_MODE.upper()}\n")
        buf.write(f"  • Порог сходства: {FUZZY_THRESHOLD}\n")
        buf.write(f"  • Окно времени:   ±{MAX_DURATION_DIFF_SEC} сек\n")
        buf.write(
            f"  • Веса:           Имя={WEIGHT_NAME} | Время={WEIGHT_DURATION} | Размер={WEIGHT_SIZE}\n"
        )
        buf.write(
            f"  • Степени (p):    Имя={NAME_POWER} | Время={DURATION_POWER} | Размер={SIZE_POWER}\n"
        )
        buf.write(f"  • Штраф (числа):  {PENALTY_NUMBERS_MISMATCH}\n")
        buf.write(f"  • Мера Жаккара:   {'ВКЛ' if USE_JACCARD_PENALTY else 'ВЫКЛ'}\n")
        buf.write(f"  • Meta fuzzy:     {'ВКЛ' if USE_META_FUZZY else 'ВЫКЛ'}\n")
        buf.write("  • Связи:          score=итог | текст/длит/размер=вклад(0..1) | штраф\n")
    else:
        buf.write("\n[НАСТРОЙКИ ПОИСКА]\n")
        buf.write("  • Режим:          STRICT (Точное совпадение)\n")

    buf.write("=" * 60 + "\n\n")

    for i, group in enumerate(groups, 1):
        sorted_group = _order_group_by_keep_priority(group)

        buf.write(f"--- ГРУППА #{i} (Файлов: {len(group)}) ---\n")

        for pos, row in enumerate(sorted_group):
            file_name = row["file_name"] or "Без названия"
            performer = (row["performer"] or "").strip()
            title = (row["title"] or "").strip()
            track_meta = " — ".join(x for x in (performer, title) if x) or "не указано"

            size_mb = _format_bytes(row["file_size"] or 0)
            msg_id = row["message_id"]
            msg_uid = row["file_unique_id"]
            dur_str = _format_duration(row["duration"])
            link = f"https://t.me/c/{clean_chat_id}/{msg_id}"

            marker = "[KEEP] " if pos == 0 else ""
            buf.write(f"• {marker}{file_name}\n")
            buf.write(f"  Track: {track_meta}\n")
            buf.write(f"  Info: {size_mb} | Время: {dur_str} | ID: {msg_id} | UID: {msg_uid}\n")
            buf.write(f"  Link: {link}\n")
            buf.write("\n")

        group_edges = edges_by_group.get(i - 1, [])
        if group_edges:
            group_edges.sort(key=lambda e: e[2].score, reverse=True)
            buf.write(f"  Связи ({len(group_edges)}):\n")
            for a, b, info in group_edges:
                if info.reason == "uid":
                    detail = "идентичный файл (UID)"
                elif info.reason == "meta":
                    detail = "точное совпадение метаданных"
                else:
                    detail = (
                        f"fuzzy: score={info.score:.3f} | "
                        f"текст={info.name:.2f}{_src_suffix(info.text_source)} | "
                        f"длит={info.dur:.2f} | размер={info.size:.2f} | "
                        f"штраф=-{info.penalty:.2f}"
                    )
                buf.write(f"    {a} <-> {b}: {detail}\n")
            buf.write("\n")

        buf.write("-" * 30 + "\n\n")

    content = buf.getvalue()
    buf.close()

    def _write() -> None:
        with open(report_file, "w", encoding="utf-8") as f:
            f.write(content)

    await asyncio.to_thread(_write)

    log.info(f"Отчет готов! Откройте файл: {report_file}")


# endregion

# region --- ОСНОВНАЯ ЛОГИКА И ВЗАИМОДЕЙСТВИЕ С TELEGRAM API ---
# Самый большой блок, содержащий всю "умную" часть скрипта: получение сообщений из Telegram, их анализ, поиск дубликатов и выполнение действий.


async def create_telegram_client() -> Client | None:
    """Создает, настраивает и возвращает экземпляр клиента Kurigram.
    В случае критической ошибки конфигурации (например, кривой прокси) возвращает None.
    """
    client_kwargs: dict[str, Any] = {
        "api_id": API_ID,
        "api_hash": API_HASH,
        "no_updates": True,
        "max_concurrent_transmissions": 10,
        "sleep_threshold": SLEEP_THRESHOLD,
        # "protocol_factory": TCPPadded, # todo в конфиг вынести
    }

    if PROXY_URL:
        if is_mtproto_link(PROXY_URL):
            try:
                local_port = await start_local_bridge(PROXY_URL)
                transport = (
                    TCPPadded if needs_padded_transport(PROXY_URL) else TCPAbridged
                )  # todo TCPIntermediatePadded
                client_kwargs["proxy"] = {
                    "scheme": "socks5",
                    "hostname": "127.0.0.1",
                    "port": local_port,
                }
                client_kwargs["protocol_factory"] = transport

                log.info(
                    f"MTProto-прокси из конфига поднят как локальный мост: "
                    f"127.0.0.1:{local_port} -> {PROXY_URL.split('server=')[-1].split('&')[0]}"
                )
            except Exception as e:
                log.critical(f"Не удалось поднять локальный мост для MTProto-прокси. Ошибка: {e}")
                return None
        else:
            try:
                parsed_proxy = urlparse(PROXY_URL)
                proxy_dict = {
                    "scheme": parsed_proxy.scheme,
                    "hostname": parsed_proxy.hostname,
                    "port": parsed_proxy.port,
                    "username": parsed_proxy.username,
                    "password": parsed_proxy.password,
                }
                client_kwargs["proxy"] = proxy_dict

                log.info(
                    f"Используется прокси: {proxy_dict['scheme']}://{proxy_dict['hostname']}:{proxy_dict['port']}"
                )
            except Exception as e:
                log.critical(
                    f"Не удалось распарсить URL прокси. Проверьте правильность ссылки в конфиге. Ошибка: {e}"
                )
                return None

    return Client(SESSION_NAME, **client_kwargs)


async def resolve_chat_identifiers(
    app: Client,
    identifiers: list[str],
    banner: str | None = "Преобразую идентификаторы чатов в числовые ID...",
) -> list[ChatID]:
    """Преобразует список идентификаторов чатов (числовые ID и @usernames)
    в список уникальных числовых ID с сохранением исходного порядка.
    Использует контролируемые конкурентные запросы к API.
    """
    if banner:
        log.info(f"\n{'=' * 20}\n{banner}")

    # 1. Дедуп строк — чтобы не слать лишние запросы к API
    unique_identifiers: list[str] = []
    seen: set[str] = set()
    for ident in identifiers:
        clean = ident.strip()
        if not clean:
            continue
        if clean not in seen:
            seen.add(clean)
            unique_identifiers.append(clean)
        else:
            log.debug(f"Пропущен дубликат во входящем списке: '{clean}'")

    semaphore = asyncio.Semaphore(VERIFY_CONCURRENCY)

    # 2. Резолв одного идентификатора: число -> само, имя -> API (или None)
    async def resolve_one(ident: str) -> ChatID | None:
        try:
            chat_id = int(ident)
            log.debug(f"Идентификатор '{ident}' распознан как числовой ID.")
            return chat_id
        except ValueError:
            log.debug(f"'{ident}' не является числом. Отправляю запрос к API.")
        async with semaphore:
            try:
                chat = await app.get_chat(ident)
            except (UsernameNotOccupied, PeerIdInvalid):
                log.error(f"Не удалось найти чат с именем '{ident}'. Он будет пропущен...")
                return None
            except UsernameInvalid:
                log.error(f"Имя пользователя '{ident}' невалидно. Оно будет пропущено...")
                return None
            except Exception as e:
                log.error(
                    f"Произошла непредвиденная ошибка при обработке '{ident}': {e}. Он будет пропущен..."
                )
                return None
        if not chat.id:
            log.error(f"Для имени пользователя '{ident}' не получен id. Будет пропущен...")
            return None
        remember_chat(chat)
        log.info(f"Имя пользователя '{ident}' успешно преобразовано в ID: {chat.id}")
        return chat.id

    # gather сохраняет порядок задач -> порядок входа сохраняется сам собой
    results = await asyncio.gather(*(resolve_one(i) for i in unique_identifiers))

    # 3. Дедуп ID — юзернейм и число могут указывать на один чат
    seen_ids: set[ChatID] = set()
    final_unique_ids: list[ChatID] = []
    duplicates_found: list[ChatID] = []
    for chat_id in results:
        if chat_id is None:
            continue
        if chat_id in seen_ids:
            duplicates_found.append(chat_id)
        else:
            seen_ids.add(chat_id)
            final_unique_ids.append(chat_id)

    if duplicates_found:
        duplicate_counts = Counter(duplicates_found)
        log.warning(
            f"В итоговом списке ID обнаружены дубликаты (возможно, юзернейм указывает на тот же ID): "
            f"{dict(duplicate_counts)}. Каждый чат будет обработан только один раз."
        )

    return final_unique_ids


async def resolve_and_validate_archive_target(app: Client, me_id: int) -> ChatID | None:
    """Резолвит ARCHIVE_TARGET в числовой ID и проверяет возможность записи.

    Возвращает ID при успехе, иначе None. 'me'/'self' → Избранное.
    Жёстко проверяются только каналы (нужны права на постинг); для групп
    полагаемся на runtime fail-safe в _archive_chunk.
    """
    log.info(f"\n{'=' * 20}\nПроверяю архивную цель '{ARCHIVE_TARGET}'...")
    resolved = await resolve_chat_identifiers(app, [ARCHIVE_TARGET], banner=None)
    if not resolved:
        log.error(f"Не удалось разрешить archive_target='{ARCHIVE_TARGET}'.")
        return None
    target_id = resolved[0]

    try:
        chat = await app.get_chat(target_id)
        remember_chat(chat)
    except Exception as e:
        log.error(f"Архивный чат '{ARCHIVE_TARGET}' ({target_id}) недоступен: {e}")
        return None

    if target_id == me_id:
        log.info("Архивная цель: Избранное (Saved Messages).")
        return target_id

    if chat.type == ChatType.CHANNEL:
        try:
            member = await app.get_chat_member(target_id, me_id)
        except UserNotParticipant:
            log.error(f"Вы не участник канала {target_id} — публикация в архив невозможна.")
            return None
        privs = getattr(member, "privileges", None)
        can_post = bool(privs and privs.can_post_messages)
        if (
            member.status not in (ChatMemberStatus.OWNER, ChatMemberStatus.ADMINISTRATOR)
            or not can_post
        ):
            log.error(f"Нет прав на публикацию в канал {target_id}.")
            return None

    log.info(
        f"Архивная цель валидна: {target_id} (режим: {ARCHIVE_MODE}, hide_sender={ARCHIVE_HIDE_SENDER})."
    )
    return target_id


async def populate_ignore_list(app: Client) -> None:
    """Обрабатывает RAW_IGNORE_LIST и RAW_IGNORE_REGEX: разрешает юзернеймы в ID
    и заполняет IGNORE_MESSAGES / IGNORE_REGEX / GLOBAL_IGNORE_REGEX.
    """
    if not RAW_IGNORE_LIST and not RAW_IGNORE_REGEX:
        return

    log.info(f"\n{'=' * 20}\nОбработка списков исключений (ignore_list / ignore_regex)...")

    # (key, payload, applier) — applier знает, куда положить данные после резолвинга
    usernames_to_resolve: list[tuple[str, Callable[[int], None]]] = []

    def _dispatch(key: str, apply: Callable[[int], None]) -> None:
        try:
            apply(int(key))
        except ValueError:
            usernames_to_resolve.append((key, apply))

    for key, msg_ids in RAW_IGNORE_LIST.items():
        _dispatch(key, lambda cid, ids=msg_ids: IGNORE_MESSAGES[cid].update(ids))

    for key, patterns in RAW_IGNORE_REGEX.items():
        if key == "*":
            GLOBAL_IGNORE_REGEX.extend(patterns)
            continue
        _dispatch(key, lambda cid, pats=patterns: IGNORE_REGEX[cid].extend(pats))

    if not usernames_to_resolve:
        log.info("Все идентификаторы в списках исключений корректны.")
        return

    log.info(f"Проверяю {len(usernames_to_resolve)} имен/ссылок...")
    semaphore = asyncio.Semaphore(VERIFY_CONCURRENCY)

    async def resolve_task(identifier: str):
        async with semaphore:
            chat = await app.get_chat(identifier)
            remember_chat(chat)
            return chat.id

    tasks = [resolve_task(k) for k, _ in usernames_to_resolve]
    results = await asyncio.gather(*tasks, return_exceptions=True)

    errors = [
        (name, r) for (name, _), r in zip(usernames_to_resolve, results) if isinstance(r, Exception)
    ]
    if errors:
        log.critical(
            "ОШИБКА КОНФИГУРАЦИИ: Не удалось проверить идентификаторы в списках исключений:"
        )
        for username, exc in errors:
            log.critical(f" ID/Name '{username}' выдало: {exc}")
        raise IgnoreListResolutionError(f"Ошибок проверки имен: {len(errors)}")

    for (_, apply), chat_id in zip(usernames_to_resolve, results):
        apply(chat_id)

    log.info(f"Успешно обработано {len(results)} юзернеймов.")


async def can_process_chat(app: Client, chat_id: ChatID, me_id: int, args: Namespace) -> bool:
    """Проверяет права доступа к чату."""
    try:
        log.info(f"\n{'=' * 20}\nПроверка прав для чата {chat_label(chat_id)}...")

        try:
            chat = await app.get_chat(chat_id)
            remember_chat(chat)
        except PeerIdInvalid:
            log.error(f"Чат {chat_label(chat_id)} не найден или недоступен.")
            return False

        # 1. Определяем режим работы
        is_read_only = args.command in ("report", "download") or DRY_RUN or REPORT_ONLY

        # 2. Личный чат — всегда разрешено
        if chat.type == ChatType.PRIVATE:
            log.info(f"Чат {chat_label(chat_id)} — личный диалог. Разрешено.")
            return True

        # 3. Проверяем членство
        try:
            member = await app.get_chat_member(chat_id, me_id)
        except UserNotParticipant:
            if is_read_only and chat.username:
                log.warning("Не участник, но чат публичный. Разрешено (read-only).")
                return True
            log.error(f"Не участник чата {chat_label(chat_id)}. Пропущен.")
            return False

        # 4. Read-only режим — права на удаление не нужны
        if is_read_only:
            log.info("Режим 'только чтение'. Разрешено.")
            return True

        # 5. Боевой режим — нужны права на удаление
        has_delete_rights = member.status == ChatMemberStatus.OWNER or (
            member.privileges and member.privileges.can_delete_messages
        )

        if has_delete_rights:
            log.info(f"Права на удаление в чате {chat_label(chat_id)} подтверждены.")
            return True

        log.error(f"Нет прав на удаление в чате {chat_label(chat_id)}. Пропущен.")
        return False

    except Exception as e:
        log.error(f"Ошибка проверки прав в чате {chat_label(chat_id)}: {e}")
        return False


def _get_audio_attributes(message: types.Message | None) -> AudioMeta | None:
    """Проверяет, является ли сообщение аудио или аудио-документом.
    Возвращает AudioMeta с атрибутами
    (file_unique_id, file_name, file_size, duration, performer, title)
    или None, если это не аудиофайл.
    """
    if not message or message.empty or message.service:
        return None

    if message.audio:
        return AudioMeta(
            file_unique_id=message.audio.file_unique_id,
            file_name=message.audio.file_name,
            file_size=message.audio.file_size,
            duration=message.audio.duration or 0,
            performer=message.audio.performer,
            title=message.audio.title,
        )

    if (
        message.document
        and message.document.mime_type
        and message.document.mime_type.startswith("audio/")
    ):
        return AudioMeta(
            file_unique_id=message.document.file_unique_id,
            file_name=message.document.file_name,
            file_size=message.document.file_size,
            duration=0,
            performer=None,
            title=None,
        )

    return None


# endregion

# region --- Под-блок: Синхронизация истории ---


async def _flush_audio_batch(
    conn: aiosqlite.Connection,
    batch: list[tuple],
) -> int:
    """Вставляет батч аудио и коммитит."""
    cur = await conn.executemany(
        "INSERT OR IGNORE INTO audios (chat_id, message_id, file_unique_id, file_name, file_size, duration, performer, title) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
        batch,
    )
    await conn.commit()
    return cur.rowcount


async def _get_media_total(
    app: Client,
    chat_id: ChatID,
    media_filter: MessagesFilter,
    is_incremental: bool,
) -> int | None:
    """Пытается получить общее количество сообщений для прогресса.
    Возвращает None если не удалось или не нужно (инкрементальный режим).
    """
    if is_incremental:
        return None
    try:
        count = await app.search_messages_count(chat_id, filter=media_filter)
        return count if count > 0 else None
    except Exception:
        return None


async def sync_messages(
    app: Client,
    chat_id: ChatID,
    conn: aiosqlite.Connection,
) -> None:
    """Синхронизация: поиск AUDIO + DOCUMENT на серверах Telegram.

    • Каждый батч коммитится отдельно.
    • Курсор обновляется ТОЛЬКО в конце.
    • INSERT OR IGNORE — идемпотентность.
    """
    log.info(f"\n{'=' * 40}\nНачинаю синхронизацию чата {chat_label(chat_id)}")

    # ── 1. Читаем состояние ──────────────────────────────────
    async with conn.execute(
        "SELECT is_fully_synced, newest_scanned_id FROM chat_sync_state WHERE chat_id = ?",
        (chat_id,),
    ) as c:
        row = await c.fetchone()

    is_fully_synced = row[0] if row else 0
    db_newest_id = row[1] if row else 0

    max_id_found = db_newest_id
    try:
        async for m in app.get_chat_history(chat_id, limit=1):
            max_id_found = max(max_id_found, m.id)
    except Exception:
        pass

    total_added = 0

    search_kwargs: dict = {}
    is_incremental = is_fully_synced and db_newest_id > 0

    if is_incremental:
        search_kwargs["min_id"] = db_newest_id
        log.info(f"Инкрементальная синхронизация (ID > {db_newest_id})")

    FILTERS_MAP = {
        MessagesFilter.AUDIO: "АУДИО",
        MessagesFilter.DOCUMENT: "ДОКУМЕНТЫ",
    }

    LOG_INTERVAL = 5.0  # секунд между строками прогресса

    # ── 2. Сканируем и пишем побатчево ───────────────────────
    try:
        for media_filter, filter_name in FILTERS_MAP.items():
            total_count = await _get_media_total(
                app,
                chat_id,
                media_filter,
                is_incremental,
            )

            batch: list[tuple] = []
            filter_added = 0
            scanned = 0
            last_log_time = time.monotonic()

            async for message in app.search_messages(
                chat_id,
                filter=media_filter,
                **search_kwargs,
            ):
                scanned += 1

                if message.id > max_id_found:
                    max_id_found = message.id

                audio_attrs = _get_audio_attributes(message)
                if audio_attrs:
                    batch.append((message.chat.id, message.id, *audio_attrs))

                # ── батч заполнен → коммитим ─────────────────
                if len(batch) >= SYNC_BATCH_SIZE:
                    added = await _flush_audio_batch(conn, batch)
                    filter_added += added
                    total_added += added
                    batch.clear()

                # ── периодический лог ────────────────────────
                now = time.monotonic()
                if now - last_log_time >= LOG_INTERVAL:
                    progress = f" / {total_count}" if total_count else ""
                    log.info(
                        f"  {filter_name}: "
                        f"просмотрено {scanned}{progress}, "
                        f"добавлено {filter_added}"
                    )
                    last_log_time = now

            # ── остаток ──────────────────────────────────────
            if batch:
                added = await _flush_audio_batch(conn, batch)
                filter_added += added
                total_added += added
                batch.clear()

            log.info(f"  {filter_name}: готово. Просмотрено {scanned}, добавлено {filter_added}")

        # ── 3. Фиксируем курсор ─────────────────────────────
        await conn.execute(
            "INSERT OR REPLACE INTO chat_sync_state "
            "(chat_id, is_fully_synced, newest_scanned_id) "
            "VALUES (?, 1, ?)",
            (chat_id, max_id_found),
        )
        await conn.commit()
        log.info(f"Синхронизация завершена. Новых записей: {total_added}")

    except Exception as e:
        await conn.rollback()
        log.error(
            f"Ошибка синхронизации чата {chat_label(chat_id)}: {e}",
            exc_info=True,
        )
        raise


# endregion

# region --- Под-блок: Поиск и обработка дубликатов ---


async def find_and_process_duplicates(
    app: Client,
    chat_id: ChatID,
    conn: aiosqlite.Connection,
    archive_target_id: ChatID | None = None,
) -> None:
    """ЭТАП 2 (Оркестратор): Анализирует дубликаты и формирует списки действий."""
    log.info(f"\n{'=' * 10}\nНачинаю анализ дубликатов в чате {chat_label(chat_id)}...")

    # Шаг 1: Найти группы потенциальных дубликатов в локальной базе данных
    potential_groups, _ = await _get_potential_duplicate_groups(chat_id, conn)
    if not potential_groups:
        log.info(f"В чате {chat_label(chat_id)} дубликатов не найдено.")
        return

    log.info(
        f"Найдено {len(potential_groups)} групп потенциальных дубликатов. Начинаю верификацию через API..."
    )

    # Шаг 2: Проверить все сообщения из этих групп, запросив их у Telegram
    ids_to_verify = list({record["message_id"] for group in potential_groups for record in group})
    verified_messages = await _verify_messages_from_api(app, chat_id, ids_to_verify)

    # Шаг 3: Классифицировать дубликаты на основе верифицированных данных
    tg_ids, db_ids, update_records = _classify_verified_duplicates(
        potential_groups, verified_messages
    )

    # Шаг 4: Передать отсортированные списки на исполнение
    await handle_database_changes(
        app,
        chat_id,
        conn,
        sorted(list(tg_ids)),
        sorted(list(db_ids)),
        update_records,
        archive_target_id=archive_target_id,
    )


def _group_audios_by_duplicates(all_audios: list[DBRow]) -> tuple[list[DuplicateGroup], EdgeMeta]:
    """(ЧИСТАЯ ФУНКЦИЯ) Группирует записи, используя обход графа
    для нахождения связных компонентов (транзитивных связей).

    Returns:
        Кортеж (groups, edge_meta): группы дубликатов (len >= 2) и
        метаданные связей с причиной "uid"/"meta" для отчёта.
    """
    if not all_audios:
        return [], {}

    id_to_record = {rec["message_id"]: rec for rec in all_audios}
    uid_to_ids = defaultdict(list)
    meta_to_ids = defaultdict(list)

    for rec in all_audios:
        msg_id = rec["message_id"]
        uid_to_ids[rec["file_unique_id"]].append(msg_id)
        meta_key = (
            rec["file_name"],
            rec["performer"],
            rec["title"],
            rec["file_size"],
            rec["duration"],
        )
        meta_to_ids[meta_key].append(msg_id)

    potential_duplicate_groups = []
    processed_ids = set()
    edge_meta: EdgeMeta = {}

    for start_msg_id in id_to_record:
        if start_msg_id in processed_ids:
            continue

        # Начинаем обход нового компонента связности
        current_group = {start_msg_id}
        work_set = {start_msg_id}  # Используем set как неупорядоченный буфер
        processed_ids.add(start_msg_id)  # Mark-on-push (сразу помечаем стартовый)

        while work_set:
            curr_id = work_set.pop()  # Порядок извлечения не важен

            rec = id_to_record[curr_id]

            neighbors_uid = uid_to_ids.get(rec["file_unique_id"], [])
            meta_key = (
                rec["file_name"],
                rec["performer"],
                rec["title"],
                rec["file_size"],
                rec["duration"],
            )
            neighbors_meta = meta_to_ids.get(meta_key, [])

            for neighbor_id, reason in itertools.chain(
                ((n, "uid") for n in neighbors_uid),
                ((n, "meta") for n in neighbors_meta),
            ):
                if neighbor_id == curr_id:
                    continue
                key = _edge_key(curr_id, neighbor_id)
                # uid имеет приоритет над meta при описании причины
                if reason == "uid" or key not in edge_meta:
                    edge_meta[key] = EdgeInfo(
                        reason=reason,
                        score=1.0,
                        name=None,
                        dur=None,
                        size=None,
                        penalty=0.0,
                    )
                if neighbor_id not in processed_ids:
                    processed_ids.add(neighbor_id)  # Mark-on-push
                    current_group.add(neighbor_id)
                    work_set.add(neighbor_id)

        if len(current_group) > 1:
            group_records = [id_to_record[gid] for gid in current_group]
            potential_duplicate_groups.append(group_records)

    return potential_duplicate_groups, edge_meta


# todo Вынести regex в конфиг
# ─────────────────────────────────────────────────────────────
# Регулярки — модульный уровень, компилируются один раз
# ─────────────────────────────────────────────────────────────

_DOMAINS = r"(?:net|com|ru|me|fm|tv|org|biz|info|cc|xyz|ua|by|kz|top|click|su|pm)"
_MEDIA_EXT = (
    r"mp3|m4a|m4b|flac|wav|ogg|ogx|wma|aac|alac|aiff|ape|opus|wv|webm|mp4|avi|wmv|mkv|flv|mov"
)
_TRASH_SITES = r"(?:muzlome|myzuka|zaycev(?:_?net)?|zvuk|muzofon|hitmo|pesni|lightaudio(?:_ru)?|ruapporangespace|mp3pulse(?:_ru)?|jamix(?:_cc)?|ipleer(?:_com)?|skysound(?:_cc)?|vk4(?:_ru)?)"

_RE_EXT = re.compile(rf"(?:\.(?:{_MEDIA_EXT}))+$", re.IGNORECASE)
_RE_BRACKETS_AD = re.compile(rf"[\[\(][^\]\)]*\b[\w-]+\.{_DOMAINS}\b[^\]\)]*[\]\)]", re.IGNORECASE)
_RE_WWW = re.compile(r"(?:www[._]|https?://|ftp://)", re.IGNORECASE)
_RE_URLS = re.compile(rf"\b[a-z0-9]+\.{_DOMAINS}\b", re.IGNORECASE)
_RE_PURE_ID = re.compile(r"^\d{1,5}[\s_\-]\d{10,}$")
_RE_TRASH_PREFIX = re.compile(rf"^{_TRASH_SITES}[_.\-\s]+", re.IGNORECASE)
_RE_NUM_PREFIX = re.compile(r"^\d{5,7}[\s_\-]+", re.IGNORECASE)
_RE_TRASH_SUFFIX = re.compile(rf"[_.\-\s]+{_TRASH_SITES}[_.\-\s]*$", re.IGNORECASE)
_RE_COPY_SUFFIX = re.compile(r"[\s_\-]*[\(\[]\d[\)\]]$")
_RE_TRASH_IDS = re.compile(r"(?<=[a-zа-яё\d])[\s_\-]+\d{7,}(?:[\s_\-]+\d{1,4})?$", re.IGNORECASE)
_RE_DIGITS = re.compile(r"\d+")
_RE_META_PLACEHOLDER = re.compile(
    r"^\s*(?:<\s*unknown\s*>|\[\s*unknown\s*\])\s*$",
    re.IGNORECASE,
)
# _RE_HASH_SUFFIX = re.compile(r"[\s_\-]+(?=[A-Z0-9]*\d)[A-Z0-9]{6}$")


# ─────────────────────────────────────────────────────────────
# Подфункции
# ─────────────────────────────────────────────────────────────


def _clean_filename(fname: str | None) -> str:
    """Нормализует имя файла для fuzzy-сравнения.

    Удаляет медиа-расширения, рекламные вставки в скобках, сайтовые
    префиксы/суффиксы/домены, суффиксы копий («(1)», «[2]») и длинные
    цифровые ID в конце. Унифицирует разделители (``_``, ``-``) в пробелы.

    Args:
        fname: Исходное имя файла или ``None``.

    Returns:
        Очищённая строка в нижнем регистре. Если после очистки ничего не
        осталось — возвращает имя без расширения (или оригинал как fallback).
    """
    if not fname:
        return ""
    original = fname
    s = fname
    s = _RE_EXT.sub("", s)
    # s = _RE_HASH_SUFFIX.sub("", s)
    s = s.lower()

    s = _RE_COPY_SUFFIX.sub(" ", s)
    if _RE_PURE_ID.match(s.strip()):
        return s.replace("_", " ").replace("-", " ").strip()
    s = _RE_NUM_PREFIX.sub(" ", s)
    s = _RE_BRACKETS_AD.sub(" ", s)
    s = _RE_WWW.sub(" ", s)
    s = _RE_TRASH_PREFIX.sub(" ", s)
    s = _RE_TRASH_SUFFIX.sub(" ", s)
    s = s.strip()
    s = s.replace("_", " ").replace("-", " ")
    s = _RE_URLS.sub(" ", s)
    s = " ".join(s.split())
    s = _RE_TRASH_IDS.sub("", s)

    cleaned = s.strip()
    if not cleaned:
        log.debug(f"'{original}' очищен в слюни")
        return _RE_EXT.sub("", original).strip() or original
    return cleaned


def _process_for_fuzzy(cleaned_name: str) -> str:
    """default_process + схлопывание пробелов — финальная форма для fuzzy."""
    return " ".join(default_process(cleaned_name).split())


# Источники текстового совпадения: имя-имя, имя-мета, мета-имя, мета-мета
_SRC_NN, _SRC_NM, _SRC_MN, _SRC_MM = 0, 1, 2, 3
_SRC_LABEL = {
    _SRC_NN: "имя-имя",
    _SRC_NM: "имя-мета",
    _SRC_MN: "мета-имя",
    _SRC_MM: "мета-мета",
}


def _clean_meta(performer: str | None, title: str | None) -> str:
    """performer+title, очищенные тем же пайплайном, что и имя файла.

    Единая нормализация важнее точечной: имя и мета должны сравниваться
    в одной форме. Плейсхолдеры вида '<unknown>' отбрасываются как
    отсутствующие значения. Пусто, если не осталось ни performer, ни title.
    """
    parts = [p for p in (performer, title) if p and not _RE_META_PLACEHOLDER.match(p)]
    if not parts:
        return ""
    return _clean_filename(" ".join(parts))


def _src_suffix(src: int | None) -> str:
    """Подпись источника для отчёта, напр. '(мета-мета)'. Пусто для uid/meta."""
    return f"({_SRC_LABEL[src]})" if src is not None else ""


def _prepare_arrays(
    sorted_rows: list[DBRow],
) -> tuple[
    np.ndarray,  # ids
    np.ndarray,  # durations
    np.ndarray,  # sizes
    list[str],  # names (очищённые)
    list[str],  # names_processed
    list[str],  # metas_processed
    np.ndarray,  # name_lengths
    list[set[int]],  # numbers_cache (числа из имени)
    list[set[int]],  # meta_numbers_cache (числа из меты)
    list[str | None],  # uids
    dict[int, DBRow],  # id_to_row
]:
    """Строит numpy-массивы и вспомогательные структуры из отсортированных строк БД.

    Все тяжёлые преобразования (очистка имён, RapidFuzz default_process,
    извлечение чисел) выполняются здесь — по одному разу на файл.

    Args:
        sorted_rows: Записи БД, отсортированные по ``duration`` (возрастание).

    Returns:
        Кортеж из одиннадцати объектов — массивы ids/durations/sizes,
        списки имён (сырых и обработанных), мета, длины имён, кэш числовых множеств,
        список UID-ов и словарь message_id → DBRow.
    """
    ids = np.array([r["message_id"] for r in sorted_rows], dtype=np.int64)
    durations = np.array([r["duration"] or 0 for r in sorted_rows], dtype=np.int32)
    sizes = np.array([r["file_size"] or 0 for r in sorted_rows], dtype=np.float64)

    names = [_clean_filename(r["file_name"]) for r in sorted_rows]
    names_processed = [_process_for_fuzzy(n) for n in names]

    if USE_META_FUZZY:
        metas = [_clean_meta(r["performer"], r["title"]) for r in sorted_rows]
    else:
        metas = [""] * len(sorted_rows)  # фича выключена -> мета пустая всюду
    metas_processed = [_process_for_fuzzy(m) for m in metas]

    name_lengths = np.array([len(n) for n in names_processed], dtype=np.int32)
    numbers_cache = [{int(x) for x in _RE_DIGITS.findall(n)} for n in names]
    meta_numbers_cache = [{int(x) for x in _RE_DIGITS.findall(m)} for m in metas]

    uids = [r["file_unique_id"] for r in sorted_rows]
    id_to_row = {r["message_id"]: r for r in sorted_rows}

    return (
        ids,
        durations,
        sizes,
        names,
        names_processed,
        metas_processed,
        name_lengths,
        numbers_cache,
        meta_numbers_cache,
        uids,
        id_to_row,
    )


def _uid_prepass(
    ids: np.ndarray,
    uids: list[str | None],
    adjacency: defaultdict[int, set[int]],
    edge_meta: EdgeMeta,
) -> int:
    """Связывает файлы с одинаковым ``file_unique_id`` до основного цикла.

    ``file_unique_id`` означает буквально один и тот же файл на серверах
    Telegram — совпадение гарантировано, fuzzy не нужен.

    Args:
        ids:       Массив message_id (int64), параллельный ``uids``.
        uids:      Список file_unique_id (может содержать ``None``).
        adjacency: Граф смежности — модифицируется на месте.
        edge_meta: Метаданные рёбер — модифицируется на месте; для каждой
                   UID-связи пишется EdgeInfo(reason="uid").

    Returns:
        Количество добавленных UID-связей.
    """
    uid_groups: defaultdict[str, list[int]] = defaultdict(list)
    for idx, uid in enumerate(uids):
        if uid:
            uid_groups[uid].append(idx)

    stats_uid_matches = 0
    for indices in uid_groups.values():
        if len(indices) < 2:
            continue
        for a, b in combinations(indices, 2):
            id_a, id_b = int(ids[a]), int(ids[b])
            adjacency[id_a].add(id_b)
            adjacency[id_b].add(id_a)
            edge_meta[_edge_key(id_a, id_b)] = EdgeInfo(
                reason="uid", score=1.0, name=None, dur=None, size=None, penalty=0.0
            )
            stats_uid_matches += 1

    return stats_uid_matches


def _compute_window_scores(
    i: int,
    window_end: int,
    durations: np.ndarray,
    sizes: np.ndarray,
    buf_thresholds: np.ndarray,
    buf_scores_dur: np.ndarray,
    buf_scores_size: np.ndarray,
    base_threshold: float,
    w_dur: float,
    w_size: float,
    dur_power: float,
    size_power: float,
) -> tuple[np.ndarray, np.ndarray, np.ndarray]:
    """Вычисляет оценки длительности и размера для окна соседей.

    Заполняет три pre-allocated буфера (views, не копии): динамические
    пороги, score по длительности и score по размеру.

    Args:
        i:                Индекс текущего файла в отсортированном массиве.
        window_end:       Правая граница скользящего окна (exclusive).
        durations:        Массив длительностей (int32).
        sizes:            Массив размеров файлов (float64).
        buf_thresholds:   Pre-allocated буфер порогов (изменяется на месте).
        buf_scores_dur:   Pre-allocated буфер score по длительности.
        buf_scores_size:  Pre-allocated буфер score по размеру.
        base_threshold:   Базовый порог схожести (FUZZY_THRESHOLD).
        w_dur:            Вес длительности (WEIGHT_DURATION).
        w_size:           Вес размера (WEIGHT_SIZE).
        dur_power:        Показатель степени для score длительности (DURATION_POWER).
        size_power:       Показатель степени для score размера (SIZE_POWER).

    Returns:
        Три view-а на буферы (dynamic_thresholds, scores_dur, scores_size)
        длиной ``window_end - i - 1``. Изменять вне функции безопасно —
        они ссылаются на те же pre-allocated массивы.
    """
    window_size = window_end - (i + 1)

    dynamic_thresholds = buf_thresholds[:window_size]
    scores_dur = buf_scores_dur[:window_size]
    scores_size = buf_scores_size[:window_size]

    dynamic_thresholds.fill(base_threshold)
    scores_dur.fill(0.0)
    scores_size.fill(0.0)

    current_dur = durations[i]
    neigh_durs = durations[i + 1 : window_end]
    curr_size = sizes[i]
    neigh_sizes = sizes[i + 1 : window_end]

    # Длительность
    valid_dur_mask = (neigh_durs > 0) & (current_dur > 0)
    if current_dur > 0 and np.any(valid_dur_mask):
        vi = np.flatnonzero(valid_dur_mask)
        v = neigh_durs[vi]
        ratio_dur = np.minimum(v / current_dur, current_dur / v)
        scores_dur[vi] = ratio_dur**dur_power

    invalid_dur_mask = ~valid_dur_mask
    if np.any(invalid_dur_mask):
        dynamic_thresholds[invalid_dur_mask] -= base_threshold * w_dur

    # Размер
    valid_size_mask = (neigh_sizes > 0) & (curr_size > 0)
    if curr_size > 0 and np.any(valid_size_mask):
        vi = np.flatnonzero(valid_size_mask)
        v = neigh_sizes[vi]
        ratio_size = np.minimum(v / curr_size, curr_size / v)
        scores_size[vi] = ratio_size**size_power

    invalid_size_mask = ~valid_size_mask
    if np.any(invalid_size_mask):
        dynamic_thresholds[invalid_size_mask] -= base_threshold * w_size

    return dynamic_thresholds, scores_dur, scores_size


def _optimistic_filter(
    i: int,
    window_end: int,
    name_lengths: np.ndarray,
    scores_dur: np.ndarray,
    scores_size: np.ndarray,
    dynamic_thresholds: np.ndarray,
    w_name: float,
    w_dur: float,
    w_size: float,
    name_power: float,
    fuzzy_mode: str,
    use_meta: bool,
) -> np.ndarray:
    """Отсекает заведомо непроходных кандидатов без вызова fuzzy.

    SORT-режим без меты: верхняя граница ratio = 2·min(L1,L2)/(L1+L2) по длинам имён.
    SET-режим ИЛИ включённая мета: граница имени = 1.0 (length-bound по имени отрезал
    бы кандидатов, совпадающих по мете).

    Args:
        i:                  Индекс текущего файла.
        window_end:         Правая граница окна (exclusive).
        name_lengths:       Массив длин обработанных имён (int32).
        scores_dur:         Score по длительности (view на буфер).
        scores_size:        Score по размеру (view на буфер).
        dynamic_thresholds: Динамические пороги (view на буфер).
        w_name:             Вес имени.
        w_dur:              Вес длительности.
        w_size:             Вес размера.
        name_power:         Степень текстового score.
        fuzzy_mode:         ``"set"`` или ``"sort"``.
        use_meta:           мета fuzzy включён

    Returns:
        Булева маска длиной ``window_end - i - 1``: ``True`` — кандидат
        проходит оптимистичную проверку.
    """
    if fuzzy_mode != "set" and not use_meta:
        curr_len = float(name_lengths[i])
        neigh_lens = name_lengths[i + 1 : window_end].astype(np.float64)
        sum_lens = neigh_lens + curr_len
        max_name_ratio = np.where(
            sum_lens > 0,
            2.0 * np.minimum(neigh_lens, curr_len) / sum_lens,
            0.0,
        )
        max_name_score = max_name_ratio**name_power
        max_potential = w_name * max_name_score + w_dur * scores_dur + w_size * scores_size
    else:
        max_potential = w_name * 1.0 + w_dur * scores_dur + w_size * scores_size

    return max_potential >= dynamic_thresholds


def _compute_penalty(
    current_numbers: set[int],
    candidate_numbers: set[int],
) -> float:
    """Вычисляет штраф за несовпадение числовых токенов в именах.

    Если оба множества пусты или равны — штраф 0. При включённом
    ``USE_JACCARD_PENALTY`` штраф пропорционален расстоянию Жаккара;
    иначе фиксированный ``PENALTY_NUMBERS_MISMATCH``.

    Args:
        current_numbers:   Числа из имени текущего файла.
        candidate_numbers: Числа из имени кандидата.

    Returns:
        Штраф в диапазоне ``[0.0, PENALTY_NUMBERS_MISMATCH]``.
    """
    if current_numbers == candidate_numbers:
        return 0.0
    if USE_JACCARD_PENALTY and (current_numbers or candidate_numbers):
        union = len(current_numbers | candidate_numbers)
        intersection = len(current_numbers & candidate_numbers)
        return PENALTY_NUMBERS_MISMATCH * (1.0 - intersection / union) if union else 0.0
    return PENALTY_NUMBERS_MISMATCH


def _filter_already_connected(
    abs_indices: np.ndarray,
    valid_indices_relative: np.ndarray,
    ids: np.ndarray,
    adjacency_i: set[int] | None,
) -> tuple[np.ndarray, np.ndarray, int]:
    """Убирает из кандидатов файлы, уже связанные с текущим.

    Пропуск уже связанных (по UID или ранее найденным fuzzy) позволяет
    избежать повторных fuzzy-сравнений, результат которых не изменит граф.

    Args:
        abs_indices:             Абсолютные индексы кандидатов.
        valid_indices_relative:  Относительные индексы кандидатов (в окне).
        ids:                     Массив message_id.
        adjacency_i:             Множество соседей текущего узла (или ``None``).

    Returns:
        Отфильтрованные ``abs_indices``, ``valid_indices_relative`` и
        количество пропущенных кандидатов.
    """
    if not adjacency_i:
        return abs_indices, valid_indices_relative, 0

    mask = np.ones(abs_indices.size, dtype=bool)
    skipped = 0
    for k, abs_idx in enumerate(abs_indices.tolist()):
        if int(ids[abs_idx]) in adjacency_i:
            mask[k] = False
            skipped += 1

    if skipped:
        abs_indices = abs_indices[mask]
        valid_indices_relative = valid_indices_relative[mask]

    return abs_indices, valid_indices_relative, skipped


def _match_batch(
    i: int,
    abs_indices: np.ndarray,
    valid_indices_relative: np.ndarray,
    ids: np.ndarray,
    names: list[str],
    names_processed: list[str],
    metas_processed: list[str],
    numbers_cache: list[set[int]],
    meta_numbers_cache: list[set[int]],
    dynamic_thresholds: np.ndarray,
    scores_dur: np.ndarray,
    scores_size: np.ndarray,
    fuzz_scorer: Any,
    w_name: float,
    w_dur: float,
    w_size: float,
    name_power: float,
    adjacency: defaultdict[int, set[int]],
    edge_meta: EdgeMeta,
) -> tuple[int, int, list[float]]:
    """Векторизованное сравнение: ``process.cdist`` для массива кандидатов.

    Все источники (имя/мета × имя/мета) считаются одним cdist. Penalty
    вычисляется лениво: Stage 1 отбирает выживших по оптимистичной оценке
    (penalty=0 -> верхняя граница итогового score, безопасно т.к. penalty >= 0),
    Stage 2 считает реальный source-aware penalty и выбирает источник только
    для выживших.

    Args:
        i:                      Индекс текущего файла.
        abs_indices:            Абсолютные индексы кандидатов.
        valid_indices_relative: Относительные индексы кандидатов.
        ids:                    Массив message_id.
        names:                  Очищённые имена (для лога).
        names_processed:        Обработанные имена (для fuzzy).
        metas_processed:        Обработанная мета performer+title (для fuzzy).
        numbers_cache:          Кэш числовых множеств из имён.
        meta_numbers_cache:     Кэш числовых множеств из меты.
        dynamic_thresholds:     Динамические пороги (view на буфер).
        scores_dur:             Score по длительности (view).
        scores_size:            Score по размеру (view).
        fuzz_scorer:            ``fuzz.token_set_ratio`` или ``token_sort_ratio``.
        w_name:                 Вес имени.
        w_dur:                  Вес длительности.
        w_size:                 Вес размера.
        name_power:             Степень текстового score.
        adjacency:              Граф смежности — модифицируется на месте.
        edge_meta:              Метаданные рёбер — модифицируется на месте;
                                для каждого совпадения пишется
                                EdgeInfo(reason="fuzzy") с коэффициентами.

    Returns:
        Кортеж ``(comparisons, matches, matched_scores)`` — счётчики для статистики.
    """
    current_name = names_processed[i]
    current_meta = metas_processed[i]
    current_numbers = numbers_cache[i]
    current_meta_numbers = meta_numbers_cache[i]
    id_i = int(ids[i])

    # ── cutoff (без изменений): penalty source-aware -> cutoff только ослабляется ──
    if w_name > 0:
        min_name_powered_scores = (
            dynamic_thresholds[valid_indices_relative]
            - scores_dur[valid_indices_relative] * w_dur
            - scores_size[valid_indices_relative] * w_size
        ) / w_name

        possible_mask = min_name_powered_scores <= 1.0
        if not np.any(possible_mask):
            return 0, 0, []
        if not np.all(possible_mask):
            abs_indices = abs_indices[possible_mask]
            valid_indices_relative = valid_indices_relative[possible_mask]

        min_raw_scores = np.maximum(0.0, min_name_powered_scores[possible_mask]) ** (
            1.0 / name_power
        )
        global_cutoff = float(np.min(min_raw_scores)) * 100.0
    else:
        global_cutoff = 0.0

    comparisons = abs_indices.size
    abs_list = abs_indices.tolist()
    n = comparisons
    candidate_names = [names_processed[idx] for idx in abs_list]
    candidate_metas = [metas_processed[idx] for idx in abs_list]
    has_candidate_meta = any(candidate_metas)
    empty_meta_mask = np.array([not m for m in candidate_metas], dtype=bool)

    rel = valid_indices_relative
    thr = dynamic_thresholds[rel]
    dur_contrib = scores_dur[rel] * w_dur
    size_contrib = scores_size[rel] * w_size

    # ── ЕДИНЫЙ cdist: строки = queries, столбцы = choices ──
    # queries:  [0]=имя тек., [1]=мета тек. (если есть)
    # choices:  [0:n]=имена кандидатов, [n:2n]=меты кандидатов (если есть)
    queries = [current_name]
    if current_meta:
        queries.append(current_meta)
    choices = candidate_names + (candidate_metas if has_candidate_meta else [])

    dist = process.cdist(
        queries,
        choices,
        scorer=fuzz_scorer,
        processor=None,
        dtype=np.float64,
        score_cutoff=global_cutoff,
        workers=1,
    )

    # Собираем матрицу источников (P, n), fuzzy 0..100, срезами из dist.
    score_rows: list[np.ndarray] = [dist[0, 0:n]]  # NN — всегда
    src_codes: list[int] = [_SRC_NN]
    if has_candidate_meta:
        score_rows.append(dist[0, n : 2 * n])  # NM
        src_codes.append(_SRC_NM)
    if current_meta:
        score_rows.append(dist[1, 0:n])  # MN
        src_codes.append(_SRC_MN)
        if has_candidate_meta:
            score_rows.append(dist[1, n : 2 * n])  # MM
            src_codes.append(_SRC_MM)

    stacked = np.vstack(score_rows)  # (P, n)
    src_arr = np.asarray(src_codes, dtype=np.int8)
    meta_src_rows = [p for p, s in enumerate(src_codes) if s in (_SRC_NM, _SRC_MM)]
    mask_phantoms = has_candidate_meta and bool(empty_meta_mask.any()) and meta_src_rows

    # ── Stage 1: оптимистичный отбор (penalty=0 -> верхняя граница) ──
    powered_stacked = (stacked / 100.0) ** name_power
    optimistic = powered_stacked * w_name  # (P, n)
    if mask_phantoms:
        for p in meta_src_rows:
            optimistic[p, empty_meta_mask] = -np.inf

    optimistic_final = optimistic.max(axis=0) + dur_contrib + size_contrib
    survive = optimistic_final >= thr
    if not np.any(survive):
        return comparisons, 0, []

    surv_idx = np.flatnonzero(survive)  # позиции в массиве кандидатов
    surv_scores = stacked[:, surv_idx]  # (P, S)

    # ── Stage 2: реальный penalty только для выживших ──
    p_count = surv_scores.shape[0]
    penalty_stacked = np.zeros((p_count, surv_idx.size), dtype=np.float64)
    for col, s in enumerate(surv_idx.tolist()):
        abs_idx = abs_list[s]
        cand_nums = numbers_cache[abs_idx]
        cand_meta_nums = meta_numbers_cache[abs_idx]
        for p, src in enumerate(src_codes):
            cur_n = current_numbers if src in (_SRC_NN, _SRC_NM) else current_meta_numbers
            cand_n = cand_nums if src in (_SRC_NN, _SRC_MN) else cand_meta_nums
            penalty_stacked[p, col] = _compute_penalty(cur_n, cand_n)

    surv_scores_powered = (surv_scores / 100.0) ** name_power
    adjusted = surv_scores_powered * w_name - penalty_stacked
    if mask_phantoms:
        surv_empty = empty_meta_mask[surv_idx]
        if surv_empty.any():
            for p in meta_src_rows:
                adjusted[p, surv_empty] = -np.inf

    best_idx = np.argmax(adjusted, axis=0)
    cols = np.arange(surv_idx.size)
    fuzzy_scores_raw = surv_scores[best_idx, cols] / 100.0
    fuzzy_scores = fuzzy_scores_raw**name_power
    penalties = penalty_stacked[best_idx, cols]
    src_per_cand = src_arr[best_idx]

    rel_surv = rel[surv_idx]
    final_scores = (
        fuzzy_scores * w_name
        + scores_dur[rel_surv] * w_dur
        + scores_size[rel_surv] * w_size
        - penalties
    )
    match_mask = final_scores >= dynamic_thresholds[rel_surv]
    matched_scores = final_scores[match_mask].tolist()
    matched_positions = np.flatnonzero(match_mask)

    if log.isEnabledFor(10):
        for k in matched_positions:
            s = int(surv_idx[k])
            abs_idx = int(abs_indices[s])
            rel_idx = int(rel[s])
            log.debug(
                f"[MATCH] Score: {final_scores[k]:.3f} (Penalty: -{penalties[k]:.2f}) | "
                f"Text: {fuzzy_scores[k]:.2f} ({_SRC_LABEL[int(src_per_cand[k])]}), "
                f"Dur: {scores_dur[rel_idx]:.2f}, Size: {scores_size[rel_idx]:.2f} | "
                f"'{names[i]}' <==> '{names[abs_idx]}'"
            )

    for k in matched_positions.tolist():
        s = int(surv_idx[k])
        abs_idx = int(abs_indices[s])
        rel_idx = int(rel[s])
        id_j = int(ids[abs_idx])
        adjacency[id_i].add(id_j)
        adjacency[id_j].add(id_i)
        edge_meta[_edge_key(id_i, id_j)] = EdgeInfo(
            reason="fuzzy",
            score=float(final_scores[k]),
            name=float(fuzzy_scores[k]),
            dur=float(scores_dur[rel_idx]),
            size=float(scores_size[rel_idx]),
            penalty=float(penalties[k]),
            text_source=int(src_per_cand[k]),
        )

    return comparisons, int(match_mask.sum()), matched_scores


def _build_groups_bfs(
    ids: np.ndarray,
    adjacency: defaultdict[int, set[int]],
    id_to_row: dict[int, DBRow],
) -> list[DuplicateGroup]:
    """Собирает связные компоненты графа смежности обходом в ширину.

    Сложность O(N + E), где N — количество файлов, E — количество рёбер.

    Args:
        ids:        Массив всех message_id (int64).
        adjacency:  Граф смежности (только узлы с хотя бы одной связью).
        id_to_row:  Словарь message_id → DBRow.

    Returns:
        Список компонент с размером ≥ 2 (одиночные файлы исключены).
    """
    groups = []
    processed = set()

    for item_id in ids.tolist():
        if item_id in processed or item_id not in adjacency:
            continue

        component: list[DBRow] = []
        queue = deque([item_id])
        processed.add(item_id)

        while queue:
            curr = queue.popleft()
            component.append(id_to_row[curr])
            for neighbor in adjacency[curr]:
                if neighbor not in processed:
                    processed.add(neighbor)
                    queue.append(neighbor)

        if len(component) > 1:
            groups.append(component)

    return groups


def _log_stats(
    count: int,
    t_prep: float,
    t_loop: float,
    t_bfs: float,
    t_total: float,
    stats_comparisons: int,
    stats_uid_matches: int,
    stats_matches: int,
    stats_skipped_connected: int,
    num_groups: int,
    match_scores: list[float],
) -> None:
    """Выводит сводную статистику fuzzy-поиска.

    Args:
        count:                   Общее число файлов.
        t_prep:                  Время подготовки данных (сек).
        t_loop:                  Время основного цикла (сек).
        t_bfs:                   Время сборки групп BFS (сек).
        t_total:                 Полное время выполнения (сек).
        stats_comparisons:       Число пар-кандидатов, дошедших до текстового этапа.
        stats_uid_matches:       Число связей, найденных через UID.
        stats_matches:           Число связей, найденных через fuzzy.
        stats_skipped_connected: Число пропущенных уже связанных пар.
        num_groups:              Число найденных групп дубликатов.
        match_scores:            Финальные score всех fuzzy-совпадений за весь прогон.
                                 Если список непустой, выводятся min/p25/median/p75/max.
                                 Пустой список допустим (например, совпадений не найдено).
    """
    if t_loop > 0 and stats_comparisons > 0:
        ops = stats_comparisons / t_loop
        ops_str = (
            f"{ops / 1_000_000:.2f}M"
            if ops >= 1_000_000
            else f"{ops / 1_000:.1f}K"
            if ops >= 1_000
            else f"{ops:.0f}"
        )
    else:
        ops_str = "N/A"

    overhead_per_file = (t_loop * 1000 / count) if count > 0 else 0
    avg_candidates = stats_comparisons / count if count > 0 else 0

    log.info(
        f"Fuzzy-поиск: {stats_comparisons:,} пар-кандидатов, "
        f"{stats_uid_matches:,} UID-связей, "
        f"{stats_matches:,} fuzzy-связей, "
        f"{num_groups} групп"
    )
    log.info(
        f"Тайминги: подготовка={t_prep:.3f}s, цикл={t_loop:.3f}s, "
        f"BFS={t_bfs:.3f}s, всего={t_total:.3f}s"
    )
    log.info(
        f"Производительность: {ops_str} pairs/sec | "
        f"Avg кандидатов/файл: {avg_candidates:.1f} | "
        f"Overhead: {overhead_per_file:.3f}ms/файл | "
        f"Пропущено (уже связаны): {stats_skipped_connected}"
    )

    if match_scores:
        arr = np.array(match_scores, dtype=np.float64)
        log.info(
            f"Score совпадений: min={arr.min():.3f}, "
            f"p25={np.percentile(arr, 25):.3f}, "
            f"median={np.median(arr):.3f}, "
            f"p75={np.percentile(arr, 75):.3f}, "
            f"max={arr.max():.3f} "
            f"(порог={FUZZY_THRESHOLD:.3f})"
        )
    if stats_comparisons > 0:
        duplicate_rate = stats_matches / stats_comparisons * 100
        log.info(f"Доля дубликатов: {duplicate_rate:.1f}%")


# ─────────────────────────────────────────────────────────────
# Оркестратор
# ─────────────────────────────────────────────────────────────


def _group_audios_fuzzy_optimized(all_audios: list[DBRow]) -> tuple[list[DuplicateGroup], EdgeMeta]:
    """Находит группы дубликатов аудиофайлов через fuzzy matching.

    Использует sliding window по отсортированным длительностям + NumPy
    для векторизации. Сравнение имён — RapidFuzz (token_set/sort_ratio).

    Args:
        all_audios: Список записей из БД с полями message_id, duration,
            file_size, file_name, file_unique_id, performer, title.

    Returns:
        Кортеж (groups, edge_meta):
          - groups: список групп, где каждая группа — list[DBRow] с len >= 2.
            Одиночные файлы (без дубликатов) не включаются.
          - edge_meta: метаданные связей (причина + коэффициенты) для отчёта.
    """
    if not all_audios:
        return [], {}

    count = len(all_audios)
    log.info(f"Запуск векторизованного Fuzzy поиска для {count} файлов...")

    t_total_start = time.perf_counter()

    # 1. Подготовка
    t_prep_start = time.perf_counter()
    sorted_rows = sorted(all_audios, key=lambda r: r["duration"] or 0)
    (
        ids,
        durations,
        sizes,
        names,
        names_processed,
        metas_processed,
        name_lengths,
        numbers_cache,
        meta_numbers_cache,
        uids,
        id_to_row,
    ) = _prepare_arrays(sorted_rows)

    BASE_THRESHOLD = FUZZY_THRESHOLD
    W_NAME = WEIGHT_NAME
    W_DUR = WEIGHT_DURATION
    W_SIZE = WEIGHT_SIZE
    MAX_DIFF = MAX_DURATION_DIFF_SEC
    NAME_PWR = NAME_POWER
    DUR_POWER = DURATION_POWER
    SZ_POWER = SIZE_POWER

    window_ends = np.searchsorted(durations, durations + MAX_DIFF, side="right")
    max_window_size = max(1, int(np.max(window_ends - np.arange(count) - 1)))
    buf_thresholds = np.empty(max_window_size, dtype=np.float64)
    buf_scores_dur = np.empty(max_window_size, dtype=np.float64)
    buf_scores_size = np.empty(max_window_size, dtype=np.float64)

    if FUZZY_MATCHING_MODE == "set":
        fuzz_scorer = fuzz.token_set_ratio
        log.info("Режим Fuzzy: SET (Агрессивный, ищет пересечения слов)")
    else:
        fuzz_scorer = fuzz.token_sort_ratio
        log.info("Режим Fuzzy: SORT (Строгий, чувствителен к разным словам)")

    adjacency: defaultdict[int, set[int]] = defaultdict(set)
    edge_meta: EdgeMeta = {}

    # 2. UID предпроход
    stats_uid_matches = _uid_prepass(ids, uids, adjacency, edge_meta)
    t_prep_end = time.perf_counter()

    # 3. Основной цикл (Sliding Window)
    t_loop_start = time.perf_counter()
    stats_comparisons = 0
    stats_matches = 0
    stats_skipped_connected = 0
    all_match_scores: list[float] = []

    match_kwargs = dict(
        ids=ids,
        names=names,
        names_processed=names_processed,
        metas_processed=metas_processed,
        numbers_cache=numbers_cache,
        meta_numbers_cache=meta_numbers_cache,
        fuzz_scorer=fuzz_scorer,
        w_name=W_NAME,
        w_dur=W_DUR,
        w_size=W_SIZE,
        name_power=NAME_PWR,
        adjacency=adjacency,
        edge_meta=edge_meta,
    )

    for i in range(count):
        window_end = window_ends[i]
        if window_end <= i + 1:
            continue

        dynamic_thresholds, scores_dur, scores_size = _compute_window_scores(
            i,
            window_end,
            durations,
            sizes,
            buf_thresholds,
            buf_scores_dur,
            buf_scores_size,
            BASE_THRESHOLD,
            W_DUR,
            W_SIZE,
            DUR_POWER,
            SZ_POWER,
        )

        candidates_mask = _optimistic_filter(
            i,
            window_end,
            name_lengths,
            scores_dur,
            scores_size,
            dynamic_thresholds,
            W_NAME,
            W_DUR,
            W_SIZE,
            NAME_PWR,
            FUZZY_MATCHING_MODE,
            USE_META_FUZZY,
        )
        if not np.any(candidates_mask):
            continue

        valid_indices_relative = np.flatnonzero(candidates_mask)
        abs_indices = valid_indices_relative + (i + 1)
        id_i = int(ids[i])

        abs_indices, valid_indices_relative, skipped = _filter_already_connected(
            abs_indices,
            valid_indices_relative,
            ids,
            adjacency.get(id_i),
        )
        stats_skipped_connected += skipped
        if abs_indices.size == 0:
            continue

        shared = dict(
            abs_indices=abs_indices,
            valid_indices_relative=valid_indices_relative,
            dynamic_thresholds=dynamic_thresholds,
            scores_dur=scores_dur,
            scores_size=scores_size,
        )

        c, m, scores = _match_batch(i, **shared, **match_kwargs)
        all_match_scores.extend(scores)

        stats_comparisons += c
        stats_matches += m

    t_loop_end = time.perf_counter()

    # 4. Сборка групп BFS
    t_bfs_start = time.perf_counter()
    groups = _build_groups_bfs(ids, adjacency, id_to_row)
    t_bfs_end = time.perf_counter()

    # 5. Статистика
    _log_stats(
        count=count,
        t_prep=t_prep_end - t_prep_start,
        t_loop=t_loop_end - t_loop_start,
        t_bfs=t_bfs_end - t_bfs_start,
        t_total=t_bfs_end - t_total_start,
        stats_comparisons=stats_comparisons,
        stats_uid_matches=stats_uid_matches,
        stats_matches=stats_matches,
        stats_skipped_connected=stats_skipped_connected,
        num_groups=len(groups),
        match_scores=all_match_scores,
    )

    return groups, edge_meta


# ─────────────────────────────────────────────────────────────
# Стратегия выбора оригинала (keep_priority)
# ─────────────────────────────────────────────────────────────


class KeepCriterion(NamedTuple):
    """extract возвращает числовое значение критерия или None (= значение
    отсутствует; такая запись проигрывает записям, у которых оно есть).
    """

    extract: Callable[[DBRow], float | None]
    prefer_max: bool


def _extract_positive(field: str) -> Callable[[DBRow], float | None]:
    def inner(r: DBRow) -> float | None:
        v = r[field]
        return float(v) if v and v > 0 else None

    return inner


_META_PLACEHOLDERS: Final[frozenset[str]] = frozenset(
    {
        "<unknown>",
        "unknown",
        "unknown artist",
        "[unknown]",
    }
)


def _meta_field_ok(value: str | None) -> bool:
    v = (value or "").strip()
    return bool(v) and v.lower() not in _META_PLACEHOLDERS


def _extract_has_meta(r: DBRow) -> float:
    """0..2: качество тегов. Плейсхолдеры и title, совпадающий с именем
    файла (мусор от сайтов-качалок), не считаются метаданными.
    """
    score = 0.0
    if _meta_field_ok(r["performer"]):
        score += 1.0
    if _meta_field_ok(r["title"]) and _clean_filename(r["title"]) != _clean_filename(
        r["file_name"]
    ):
        score += 1.0
    return score


def _extract_clean_name_len(r: DBRow) -> float | None:
    cleaned = _clean_filename(r["file_name"])
    return float(len(cleaned)) if cleaned else None


_KEEP_CRITERIA: Final[dict[str, KeepCriterion]] = {
    "oldest": KeepCriterion(lambda r: float(r["message_id"]), prefer_max=False),
    "newest": KeepCriterion(lambda r: float(r["message_id"]), prefer_max=True),
    "largest": KeepCriterion(_extract_positive("file_size"), prefer_max=True),
    "smallest": KeepCriterion(_extract_positive("file_size"), prefer_max=False),
    "longest": KeepCriterion(_extract_positive("duration"), prefer_max=True),
    "shortest": KeepCriterion(_extract_positive("duration"), prefer_max=False),
    "best_meta": KeepCriterion(_extract_has_meta, prefer_max=True),
    "longest_clean_name": KeepCriterion(_extract_clean_name_len, prefer_max=True),
}

# Ловим рассинхрон реестра и валидации конфига на импорте, а не в рантайме
assert set(_KEEP_CRITERIA) == _KEEP_CRITERIA_VALID, (
    "Реестр критериев main.py разошёлся с _KEEP_CRITERIA_VALID в config.py"
)


def _cascade_winner(pool: list[DBRow]) -> DBRow:
    """Выбирает лучшего кандидата каскадом критериев KEEP_PRIORITY.

    На каждом уровне: записи без значения отсеиваются (если у кого-то
    значение есть), затем остаются все в пределах допуска от лучшего.
    Уникальный tie-break (oldest/newest) в конце списка гарантирует,
    что каскад завершится ровно одним кандидатом.
    """
    cands = pool
    for name, tol in KEEP_PRIORITY:
        if len(cands) == 1:
            break
        crit = _KEEP_CRITERIA[name]
        scored = [(crit.extract(r), r) for r in cands]
        valid = [(v, r) for v, r in scored if v is not None]
        if not valid:
            continue  # критерий неприменим ко всей группе — пропускаем уровень

        best = max(v for v, _ in valid) if crit.prefer_max else min(v for v, _ in valid)
        eps = abs(best) * tol
        cands = [r for v, r in valid if abs(v - best) <= eps]
    return cands[0]


def _order_group_by_keep_priority(group: DuplicateGroup) -> list[DBRow]:
    """Полный порядок приоритета: [оригинал, fallback #1, fallback #2, ...].

    Порядок нужен целиком: если лучший кандидат не пройдёт верификацию
    (удалён/изменён), оригиналом станет следующий. Повторный каскад по
    остатку — O(n²·C), но группы дубликатов крошечные.
    """
    pool = list(group)
    ordered: list[DBRow] = []
    while pool:
        winner = _cascade_winner(pool)
        ordered.append(winner)
        pool = [r for r in pool if r is not winner]
    return ordered


async def _get_potential_duplicate_groups(
    chat_id: ChatID, conn: aiosqlite.Connection
) -> tuple[list[DuplicateGroup], EdgeMeta]:
    """ЭТАП 2.1: Запрашивает из БД все аудио и передает их чистой функции для группировки.
    Возвращает группы дубликатов и метаданные связей (причина + коэффициенты)
    для отчёта.
    """
    log.debug(f"Анализ чата {chat_label(chat_id)}: Запрос всех аудиозаписей из локальной БД...")
    # Лимит тележки на сообщения 1 млн, т.е. на 1 чат максимум 1 млн записей, что не бьёт по ОЗУ
    # Но фактически маловероятно, что есть чаты, где аудиофайлов больше, чем 300 тыс., что максимум для 512 МБ - 1 ГБ ОЗУ
    # Т.е. здесь сделано верно
    async with conn.execute("SELECT * FROM audios WHERE chat_id = ?", (chat_id,)) as cursor:
        all_audios = await cursor.fetchall()

    if not all_audios:
        log.info(f"В базе данных нет записей для анализа в чате {chat_label(chat_id)}.")
        return [], {}

    if ENABLE_FUZZY_MATCHING:
        return await asyncio.to_thread(_group_audios_fuzzy_optimized, all_audios)
        # from bench_fuzzy_crossover import bench_fuzzy_crossover
        # await asyncio.to_thread(bench_fuzzy_crossover, all_audios)
        # return []
    else:
        return await asyncio.to_thread(_group_audios_by_duplicates, all_audios)


async def _verify_messages_from_api(
    app: Client, chat_id: ChatID, ids_to_verify: list[MessageID]
) -> VerifiedMessagesDict:
    """ЭТАП 2.2: Надежно запрашивает у Telegram информацию о сообщениях по их ID.
    Использует семафор для контроля параллельных запросов и пакетирование.
    """
    verified_messages = {}
    semaphore = asyncio.Semaphore(VERIFY_CONCURRENCY)

    async def fetch_chunk(chunk_ids):
        async with semaphore:
            try:
                return await app.get_messages(chat_id, chunk_ids)
            except Exception as e:
                log.error(f"Ошибка при получении пакета сообщений (ID: {chunk_ids[0]}...): {e}.")
                return e

    original_chunks = [list(chunk) for chunk in itertools.batched(ids_to_verify, VERIFY_CHUNK_SIZE)]
    tasks = [fetch_chunk(chunk) for chunk in original_chunks]
    results_from_gather = await asyncio.gather(*tasks)

    for original_chunk, result_chunk in zip(original_chunks, results_from_gather, strict=True):
        if isinstance(result_chunk, Exception):
            for msg_id in original_chunk:
                verified_messages[msg_id] = result_chunk
        else:
            found_messages = {msg.id: msg for msg in result_chunk if msg}
            for msg_id in original_chunk:
                verified_messages[msg_id] = found_messages.get(msg_id)

    return verified_messages


def _classify_verified_duplicates(
    duplicate_groups: list[DuplicateGroup], verified_messages: VerifiedMessagesDict
) -> ClassificationResult:
    """ЭТАП 2.3: Анализирует верифицированные сообщения и принимает решение о действиях.

    ВАЖНО: Использует стратегию Fail-Safe. Если при проверке любого сообщения
    в группе возникает ошибка API, вся группа пропускается для предотвращения
    случайного удаления данных.
    """
    to_delete_from_tg = set()
    to_delete_from_db = set()
    to_update_in_db = []

    for group in duplicate_groups:
        sorted_group = _order_group_by_keep_priority(group)

        found_a_valid_original = False
        group_is_safe_to_process = True

        # --- Шаг 1: Проверка на ошибки API (Safety Check) ---
        group_msg_ids = [r["message_id"] for r in sorted_group]

        for db_record in sorted_group:
            msg_id = db_record["message_id"]
            api_result = verified_messages.get(msg_id)

            if isinstance(api_result, Exception):
                log.warning(
                    f"Не удалось проверить сообщение {msg_id} "
                    f"({type(api_result).__name__}). "
                    f"Группа дубликатов {group_msg_ids} пропущена "
                    f"во избежание потери данных."
                )
                group_is_safe_to_process = False
                break

        if not group_is_safe_to_process:
            continue

        # --- Шаг 2: Основная логика ---
        for db_record in sorted_group:
            msg_id = db_record["message_id"]
            api_result = verified_messages.get(msg_id)

            if api_result is None or api_result.empty:
                log.debug(f"Сообщение {msg_id} не найдено в Telegram. Будет удалено из БД.")
                to_delete_from_db.add(msg_id)
            else:
                api_audio_attrs = _get_audio_attributes(api_result)
                if api_audio_attrs:
                    # Проверка на изменение контента
                    if (
                        api_audio_attrs.file_unique_id == db_record["file_unique_id"]
                        and api_audio_attrs.file_name == db_record["file_name"]
                        and api_audio_attrs.file_size == db_record["file_size"]
                        and api_audio_attrs.duration == db_record["duration"]
                        and api_audio_attrs.performer == db_record["performer"]
                        and api_audio_attrs.title == db_record["title"]
                    ):
                        if not found_a_valid_original:
                            found_a_valid_original = True
                            log.debug(
                                f"Сообщение {msg_id} ('{db_record['file_name']}') является валидным оригиналом."
                            )
                        else:
                            to_delete_from_tg.add(msg_id)
                    else:
                        log.info(
                            f"Контент сообщения {msg_id} изменился. Запись в БД будет обновлена."
                        )
                        to_update_in_db.append(api_result)
                else:
                    log.info(f"Сообщение {msg_id} больше не аудио. Запись будет удалена из БД.")
                    to_delete_from_db.add(msg_id)

    return ClassificationResult(to_delete_from_tg, to_delete_from_db, to_update_in_db)


# endregion

# region --- Под-блок: Применение изменений ---


async def _get_regex_protected_ids(
    conn: aiosqlite.Connection,
    chat_id: ChatID,
    tg_ids: list[MessageID],
    patterns: list[re.Pattern[str]],
) -> set[MessageID]:
    """Возвращает ID сообщений, чьи file_name/performer/title матчатся
    хотя бы одним паттерном. Читает метаданные из локальной БД.
    """
    protected: set[MessageID] = set()
    CHUNK = 4000

    for i in range(0, len(tg_ids), CHUNK):
        chunk = tg_ids[i : i + CHUNK]
        placeholders = ",".join("?" * len(chunk))
        rows = await conn.execute_fetchall(
            f"SELECT message_id, file_name, performer, title "
            f"FROM audios WHERE chat_id = ? AND message_id IN ({placeholders})",
            (chat_id, *chunk),
        )
        for msg_id, *fields in rows:
            hit = next(
                ((p, f) for f in fields if f for p in patterns if p.search(f)),
                None,
            )
            if hit:
                protected.add(msg_id)
                log.info(
                    f"Regex-защита: сообщение {msg_id} ('{hit[1]}') совпало с '{hit[0].pattern}'."
                )
    return protected


async def _filter_ignored_ids(
    conn: aiosqlite.Connection,
    chat_id: ChatID,
    tg_ids: list[MessageID],
) -> list[MessageID]:
    """Отсекает сообщения из ignore-листа и regex-защиты. Без побочных эффектов в БД."""
    ignore_list = IGNORE_MESSAGES.get(chat_id, set())
    final = [msg_id for msg_id in tg_ids if msg_id not in ignore_list]
    skipped = len(tg_ids) - len(final)
    if skipped:
        log.info(
            f"Пропускаю удаление {skipped} сообщений из чата {chat_label(chat_id)} (в списке игнорирования)."
        )

    patterns = IGNORE_REGEX.get(chat_id, []) + GLOBAL_IGNORE_REGEX
    if patterns and final:
        protected = await _get_regex_protected_ids(conn, chat_id, final, patterns)
        if protected:
            final = [msg_id for msg_id in final if msg_id not in protected]
            log.info(
                f"Пропускаю удаление {len(protected)} сообщений из чата {chat_label(chat_id)} (regex-защита)."
            )

    return final


def _log_planned_changes(
    chat_id: ChatID,
    tg_ids: list[MessageID],
    db_delete_ids: list[MessageID],
    db_update_records: list[types.Message],
    archive_target_id: ChatID | None,
) -> None:
    """DRY_RUN: печатает план без побочных эффектов."""
    log.info("РЕЖИМ СИМУЛЯЦИИ АКТИВЕН. Никаких реальных изменений не будет.")
    if tg_ids:
        if ARCHIVE_BEFORE_DELETE and archive_target_id is not None:
            log.info(
                f"Планируется {ARCHIVE_MODE} {len(tg_ids)} сообщений из чата {chat_label(chat_id)} "
                f"в архив {archive_target_id} перед удалением."
            )
        log.info(
            f"Планируется к удалению из Telegram в чате {chat_label(chat_id)} ({len(tg_ids)} шт.): {tg_ids}"
        )
    if db_delete_ids:
        log.info(
            f"Планируется к удалению из БД для чата {chat_label(chat_id)} ({len(db_delete_ids)} шт.): {db_delete_ids}"
        )
    if db_update_records:
        update_info = []
        for msg in db_update_records:
            attrs = _get_audio_attributes(msg)
            name = attrs.file_name if attrs else "<not-audio>"
            update_info.append(f"{msg.id} -> '{name}'")
        log.info(
            f"Планируется к обновлению в БД для чата {chat_label(chat_id)} ({len(db_update_records)} шт.): {update_info}"
        )


async def _send_archive_header(
    app: Client, archive_target_id: ChatID, chat_id: ChatID, count: int
) -> None:
    """Шлёт в архив текстовый разделитель перед пересылкой батчей чата.

    Косметика: падение заголовка не должно останавливать архивацию.
    """
    # title = ""
    try:
        chat = await app.get_chat(chat_id)
        remember_chat(chat)
        # title = (chat.title or getattr(chat, "first_name", "") or "").strip()
    except Exception:
        pass

    ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    text = f"Удалено из чата {chat_label(chat_id)}\nФайлов (план): {count}\n{ts}"
    try:
        await app.send_message(archive_target_id, text)
    except Exception as e:
        log.warning(f"Не удалось отправить заголовок архива для чата {chat_label(chat_id)}: {e}")


async def _archive_chunk(
    app: Client, chat_id: ChatID, archive_target_id: ChatID, chunk: list[MessageID]
) -> bool:
    """Архивирует батч (forward или copy). True — если ВЕСЬ батч успешно заархивирован.

    hide_sender_name применяется только при forward (copy и так без автора).
    """
    try:
        if ARCHIVE_MODE == "copy":
            archived = 0
            for msg_id in chunk:
                await app.copy_message(
                    chat_id=archive_target_id, from_chat_id=chat_id, message_id=msg_id
                )
                archived += 1
        else:
            result = await app.forward_messages(
                chat_id=archive_target_id,
                from_chat_id=chat_id,
                message_ids=chunk,
                hide_sender_name=ARCHIVE_HIDE_SENDER,
            )
            archived = len(result) if isinstance(result, list) else 1
    except Exception as e:
        log.error(f"Архивация батча из чата {chat_label(chat_id)} провалена: {e}")
        return False

    if archived != len(chunk):
        log.warning(
            f"Заархивировано лишь {archived}/{len(chunk)} из чата {chat_label(chat_id)}; батч не будет удалён."
        )
        return False
    return True


async def _archive_and_delete_messages(
    app: Client,
    chat_id: ChatID,
    conn: aiosqlite.Connection,
    tg_ids: list[MessageID],
    archive_target_id: ChatID | None,
) -> None:
    """Архивирует (опц.) и пакетно удаляет сообщения из Telegram и БД.

    Инвариант: батч удаляется только после успешной архивации (когда она включена
    и ABORT_DELETE_ON_ARCHIVE_FAILURE=True). Из БД чистятся только реально
    исчезнувшие из TG записи. Не коммитит — коммит в оркестраторе.
    """
    if not tg_ids:
        return

    archive_enabled = ARCHIVE_BEFORE_DELETE and archive_target_id is not None
    if archive_enabled and archive_target_id == chat_id:
        log.error(f"Архивная цель совпадает с чатом {chat_label(chat_id)} — архивация невозможна.")
        if ABORT_DELETE_ON_ARCHIVE_FAILURE:
            log.error("Удаление в этом чате пропущено во избежание потери данных.")
            return
        archive_enabled = False

    log.info(
        f"Обработка {len(tg_ids)} дубликатов в чате {chat_label(chat_id)} "
        f"(архивация: {'вкл' if archive_enabled else 'выкл'})..."
    )

    if archive_enabled:
        await _send_archive_header(app, archive_target_id, chat_id, len(tg_ids))

    batches = list(itertools.batched(tg_ids, BATCH_DELETE_SIZE))
    total = len(batches)

    for i, chunk_tuple in enumerate(batches):
        batch_num = i + 1
        chunk = list(chunk_tuple)
        log.info(f"Батч {batch_num}/{total} ({len(chunk)} сообщений)...")

        if archive_enabled:
            ok = await _archive_chunk(app, chat_id, archive_target_id, chunk)
            if not ok and ABORT_DELETE_ON_ARCHIVE_FAILURE:
                log.error(f"Батч {batch_num}/{total} не заархивирован — удаление пропущено.")
                continue

        try:
            deleted = await app.delete_messages(chat_id, chunk, revoke=REVOKE_PRIVATE_CHATS)
            deleted_count = deleted if isinstance(deleted, int) else len(chunk)

            if deleted_count != len(chunk):
                log.warning(
                    f"Батч {batch_num}/{total}: Telegram удалил {deleted_count}/{len(chunk)}; "
                    f"уточняю, какие записи реально исчезли..."
                )
                check = await app.get_messages(chat_id, chunk)
                check = check if isinstance(check, list) else [check]
                alive = {m.id for m in check if m and not m.empty}
                gone = [m_id for m_id in chunk if m_id not in alive]
            else:
                gone = chunk

            await conn.executemany(
                "DELETE FROM audios WHERE chat_id = ? AND message_id = ?",
                [(chat_id, mid) for mid in gone],
            )
            log.info(
                f"Батч {batch_num}/{total}: удалено {deleted_count}, очищено записей БД: {len(gone)}."
            )
        except Exception as e:
            log.error(
                f"Не удалось удалить батч {batch_num}/{total} в чате {chat_label(chat_id)}: {e}"
            )


async def _delete_db_records(
    conn: aiosqlite.Connection, chat_id: ChatID, db_delete_ids: list[MessageID]
) -> None:
    """Удаляет устаревшие записи из БД (сообщений уже нет в TG). Не коммитит."""
    if not db_delete_ids:
        return
    log.info(
        f"Чистка {len(db_delete_ids)} устаревших записей из БД для чата {chat_label(chat_id)}..."
    )
    await conn.executemany(
        "DELETE FROM audios WHERE chat_id = ? AND message_id = ?",
        [(chat_id, mid) for mid in db_delete_ids],
    )
    log.info(f"Удалено {len(db_delete_ids)} устаревших записей для чата {chat_label(chat_id)}.")


async def _apply_db_updates(
    conn: aiosqlite.Connection, chat_id: ChatID, db_update_records: list[types.Message]
) -> None:
    """Обновляет изменившиеся записи в БД. Не коммитит."""
    if not db_update_records:
        return
    log.info(
        f"Обновление {len(db_update_records)} изменённых записей в БД для чата {chat_label(chat_id)}..."
    )
    update_data = []
    for r in db_update_records:
        attrs = _get_audio_attributes(r)
        if attrs:
            update_data.append((*attrs, r.chat.id, r.id))
    await conn.executemany(
        "UPDATE audios SET file_unique_id=?, file_name=?, file_size=?, duration=?, "
        "performer=?, title=? WHERE chat_id=? AND message_id=?",
        update_data,
    )
    log.info(f"Обновлено {len(db_update_records)} записей для чата {chat_label(chat_id)}.")


async def handle_database_changes(
    app: Client,
    chat_id: ChatID,
    conn: aiosqlite.Connection,
    tg_ids: list[MessageID],
    db_delete_ids: list[MessageID],
    db_update_records: list[types.Message],
    archive_target_id: ChatID | None = None,
) -> None:
    """ЭТАП 3 (оркестратор): применяет изменения единой транзакцией.

    Под-функции НЕ коммитят — коммит делается здесь один раз.
    """
    if not any([tg_ids, db_delete_ids, db_update_records]):
        log.debug(f"Для чата {chat_label(chat_id)} нет запланированных изменений.")
        return

    log.info(f"\n{'=' * 10}\nПриведение в исполнение плана для чата {chat_label(chat_id)}...")

    final_tg_ids = await _filter_ignored_ids(conn, chat_id, tg_ids)

    if DRY_RUN:
        _log_planned_changes(
            chat_id, final_tg_ids, db_delete_ids, db_update_records, archive_target_id
        )
        return

    await _archive_and_delete_messages(app, chat_id, conn, final_tg_ids, archive_target_id)
    await _delete_db_records(conn, chat_id, db_delete_ids)
    await _apply_db_updates(conn, chat_id, db_update_records)

    await conn.commit()


# endregion

# region --- ГЛАВНАЯ УПРАВЛЯЮЩАЯ ЛОГИКА ---
# Финальный блок, содержащий только главную функцию main, которая выступает в роли "дирижёра" — вызывает функции из других блоков в правильном порядке.


async def process_single_chat(
    app: Client,
    chat_id: ChatID,
    me_id: int,
    args: Namespace,
    run_ts: str | None = None,
    archive_target_id: ChatID | None = None,
) -> None:
    """Полный цикл обработки одного чата (синхронизация, отчеты, удаление дубликатов)."""
    if not await can_process_chat(app, chat_id, me_id, args):
        return

    try:
        # NOTE: Подключение пересоздаётся на каждый чат намеренно —
        # изоляция по чатам важнее экономии ~2мс на PRAGMA.
        async with create_connection() as conn:
            await sync_messages(app, chat_id, conn)

            if args.command == "report" or REPORT_ONLY:
                await create_duplicates_report(chat_id, conn, ts=run_ts)
                return

            async with conn.execute(
                "SELECT is_fully_synced FROM chat_sync_state WHERE chat_id = ?", (chat_id,)
            ) as cursor:
                sync_state = await cursor.fetchone()

            is_fully_synced = sync_state and sync_state[0]

            if is_fully_synced:
                await find_and_process_duplicates(app, chat_id, conn, archive_target_id)
            else:
                log.info(
                    f"Чат {chat_label(chat_id)} еще не полностью синхронизирован. Пропуск этапа очистки дубликатов до завершения синхронизации."
                )

    except Exception as e:
        log.critical(
            f"Произошла невосстановимая ошибка при обработке чата {chat_label(chat_id)}: {e}",
            exc_info=True,
        )


# todo рефакторинг?
# todo поддержка музыки из профиля (отдельная таблица)
async def main() -> None:
    """Главная управляющая функция."""
    args = parse_arguments()

    # Экспорты
    if args.command == "export":
        if args.export_command == "filenames":
            await export_filenames_to_txt(args.chat)
            log.info("Задача экспорта имен файлов завершена. Выход.")
            return
        elif args.export_command == "filenames-url":
            await export_filenames_with_url_to_txt(args.chat)
            log.info("Задача экспорта имен файлов со ссылками завершена. Выход.")
            return
        elif args.export_command == "cleaned-names":
            await export_cleaned_names_to_csv(args.chat)
            log.info("Задача экспорта очищенных имен завершена. Выход.")
            return
        elif args.export_command == "cleaned-meta":
            await export_cleaned_meta_to_csv(args.chat)
            log.info("Задача экспорта очищенных метаданных завершена. Выход.")
            return
        elif args.export_command == "xlsx":
            await export_database_to_xlsx(args.chat)
            log.info("Задача экспорта Excel завершена. Выход.")
            return

    async with async_ipc_lock(LOCK_FILE, timeout=LOCK_TIMEOUT):
        if not await check_disk_space():
            log.critical("Работа скрипта прервана из-за недостатка свободного места.")
            return

        if args.command != "repair" and BACKUP_ON_STARTUP:
            await create_database_backup()

        app = await create_telegram_client()
        if app is None:
            return

        if args.command == "repair":
            log.info("=" * 15 + "ЗАПУСК В РЕЖИМЕ РЕМОНТА БД" + "=" * 15)
            try:
                async with app:
                    await repair_database(app)
            except Exception as e:
                log.critical(f"Произошла критическая ошибка в режиме ремонта: {e}", exc_info=True)
            return

        log.debug(f"\n{'=' * 20}")
        if DRY_RUN:
            log.warning("Скрипт запущен в режиме симуляции (dry_run = True).")
        pretty = ", ".join(f"{n} ~{t:.0%}" if t else n for n, t in KEEP_PRIORITY)
        log.info(f"Стратегия выбора оригинала: {pretty}")

        await initialize_database()

        if not await validate_database():
            log.critical("Скрипт остановлен из-за критических ошибок валидации БД.")
            return

        async with app:
            me = app.me

            if args.command == "download":
                resolved_ids = await resolve_chat_identifiers(app, [args.chat])
                if not resolved_ids:
                    log.error(f"Не удалось найти чат по идентификатору: {args.chat}")
                    return

                target_chat_id = resolved_ids[0]
                if len(resolved_ids) > 1:
                    log.info("Список обрезан до первого элемента")

                if not await can_process_chat(app, target_chat_id, me.id, args):
                    log.error("Нет доступа к чату или чат не найден.")
                    return

                await download_chat_audio(app, target_chat_id)
                log.info("Работа завершена. Выход.")
                return

            resolved_chat_list = await resolve_chat_identifiers(app, CHAT_LIST)

            try:
                await populate_ignore_list(app)
            except IgnoreListResolutionError:
                return

            archive_target_id: ChatID | None = None
            if ARCHIVE_BEFORE_DELETE and not (args.command == "report" or REPORT_ONLY):
                archive_target_id = await resolve_and_validate_archive_target(app, me.id)
                if archive_target_id is None and not DRY_RUN:
                    log.critical(
                        "Архивация включена, но целевой чат недоступен. "
                        "Останавливаюсь, чтобы не удалять без резервной копии."
                    )
                    return
                if archive_target_id in resolved_chat_list:
                    log.warning(
                        f"Архивный чат {archive_target_id} есть в chat_list — "
                        f"на следующем прогоне его содержимое может быть задедуплено."
                    )

            # Единый таймстемп прогона: файлы отчётов всех чатов получат общий
            # префикс, что позволяет собрать их как снапшот одного запуска.
            run_ts = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")

            for chat_id in resolved_chat_list:
                await process_single_chat(
                    app,
                    chat_id,
                    me.id,
                    args,
                    run_ts=run_ts,
                    archive_target_id=archive_target_id,
                )

    log.info("Работа скрипта завершена.")


if __name__ == "__main__":
    log.info(f"\n\n{('=+' * 60 + '\n') * 2}")
    if sys.platform in ("win32", "cygwin"):
        try:
            import winloop

            winloop.install()
            log.debug(f"winloop установлен как основной цикл событий (Platform: {sys.platform}).")
        except ImportError:
            log.warning(
                "winloop не найден. Рекомендуется 'pip install winloop' для ускорения на Windows."
            )
            log.debug("Используется стандартный цикл событий asyncio.")
    else:
        # Linux, macOS, BSD, и др.
        try:
            import uvloop

            uvloop.install()
            log.debug(f"uvloop установлен как основной цикл событий (Platform: {sys.platform}).")
        except ImportError:
            log.warning("uvloop не найден. Рекомендуется 'pip install uvloop' для ускорения.")
            log.debug("Используется стандартный цикл событий asyncio.")
    try:
        with secure_umask(0o077):
            asyncio.run(main())
    except AlreadyRunningError as e:
        log.warning(str(e))
        sys.exit(1)
    except Exception as e:
        log.critical(f"Критическая ошибка при выполнении скрипта: {e}", exc_info=True)

# endregion
