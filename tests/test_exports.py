"""Тесты экспортов (dedup.exports): пути, txt/csv/xlsx на синтетической БД."""

import csv
from pathlib import Path

from helpers import CHAT_ID

from dedup.exports import (
    build_export_path,
    export_cleaned_meta_to_csv,
    export_cleaned_names_to_csv,
    export_database_to_xlsx,
    export_filenames_to_txt,
    export_filenames_with_url_to_txt,
)


def exports_root() -> Path:
    from dedup.context import get_settings

    return Path(get_settings().paths.exports_dir)


def test_build_export_path_for_chat(configure_settings):
    path = build_export_path(CHAT_ID, "filenames", "txt", ts="2026-01-01_00-00-00")
    assert path == exports_root() / str(CHAT_ID) / "2026-01-01_00-00-00_filenames.txt"
    assert path.parent.is_dir()


def test_build_export_path_full_export_uses_db_stem(configure_settings):
    path = build_export_path(0, "cleaned_names", "csv", ts="2026-01-01_00-00-00")
    # db_file фабрики — test.sqlite -> стем попадает в имя
    assert path == exports_root() / "_full" / "2026-01-01_00-00-00_cleaned_names_of_test.csv"


async def test_export_filenames_to_txt(seeded_db):
    await export_filenames_to_txt(CHAT_ID)
    files = list((exports_root() / str(CHAT_ID)).glob("*_filenames.txt"))
    assert len(files) == 1
    lines = files[0].read_text(encoding="utf-8").splitlines()
    # Все строки сид-датасета кроме row 8 (file_name IS NULL) и row 12 (другой чат)
    assert len(lines) == 10


async def test_export_filenames_with_url(seeded_db):
    await export_filenames_with_url_to_txt(CHAT_ID)
    file = next((exports_root() / str(CHAT_ID)).glob("*_filenames_with_urls.txt"))
    content = file.read_text(encoding="utf-8")
    # Ссылка построена из chat_id без префикса -100
    assert "https://t.me/c/1234567890/1" in content
    assert "https://t.me/c/1234567890/11" in content


async def test_export_cleaned_names_csv(seeded_db):
    await export_cleaned_names_to_csv(CHAT_ID)
    file = next((exports_root() / str(CHAT_ID)).glob("*_cleaned_names.csv"))
    raw = file.read_bytes()
    assert raw.startswith(b"\xef\xbb\xbf")  # utf-8-sig BOM для Excel
    with open(file, encoding="utf-8-sig", newline="") as f:
        rows = list(csv.reader(f, delimiter=";"))
    assert rows[0] == ["Исходное имя", "После clean_filename", "После default_process"]
    assert len(rows) == 11  # заголовок + 10 именованных строк чата
    # Очистка применена: (1)-суффикс исчез
    cleaned_column = {row[1] for row in rows[1:]}
    assert "artist track one" in cleaned_column


async def test_export_cleaned_meta_csv(seeded_db):
    await export_cleaned_meta_to_csv(CHAT_ID)
    file = next((exports_root() / str(CHAT_ID)).glob("*_cleaned_meta.csv"))
    with open(file, encoding="utf-8-sig", newline="") as f:
        rows = list(csv.reader(f, delimiter=";"))
    # Строки с валидной метой: 1,2,3,4,5,8,11 (у 9-й только <unknown> — отфильтрована)
    assert len(rows) == 8  # заголовок + 7
    assert rows[0][0] == "Performer"


async def test_full_export_covers_all_chats(seeded_db):
    await export_cleaned_names_to_csv(0)
    file = next((exports_root() / "_full").glob("*_cleaned_names_of_test.csv"))
    with open(file, encoding="utf-8-sig", newline="") as f:
        rows = list(csv.reader(f, delimiter=";"))
    # 11 именованных строк: 10 текущего чата + строка другого чата
    assert len(rows) == 12


async def test_export_xlsx(seeded_db):
    await export_database_to_xlsx(CHAT_ID)
    file = next((exports_root() / str(CHAT_ID)).glob("*_database_export.xlsx"))

    from openpyxl import load_workbook

    wb = load_workbook(file, read_only=True)
    assert "audios" in wb.sheetnames
    ws = wb["audios"]
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0] == (
        "chat_id",
        "message_id",
        "file_unique_id",
        "file_name",
        "file_size",
        "duration",
        "performer",
        "title",
    )
    assert len(rows) == 12  # заголовок + 11 строк чата
    wb.close()


async def test_export_empty_chat_creates_no_file(fresh_db):
    await export_filenames_to_txt(-1000000000005)
    sub = exports_root() / "-1000000000005"
    assert sub.is_dir()  # каталог создаётся при построении пути
    assert list(sub.glob("*.txt")) == []


async def test_export_missing_db_is_safe(tmp_path, configure_settings):
    # db_file указывает на несуществующий файл — экспорт молча завершается
    configure_settings(paths={"db_file": str(tmp_path / "missing.sqlite")})
    await export_filenames_to_txt(CHAT_ID)
    if exports_root().exists():
        assert list(exports_root().rglob("*.txt")) == []
