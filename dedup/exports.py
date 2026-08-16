"""Экспорт данных из БД в txt/csv/xlsx."""

import asyncio
import datetime
import os
from contextlib import suppress
from pathlib import Path
from typing import Any, Final

import aiosqlite

from .config import DB_FILE, EXPORTS_DIR
from .fuzzy import _clean_filename, _clean_meta, _process_for_fuzzy
from .logger import log
from .state import chat_label
from .typedefs import ChatID, CsvRowFormatter, DBRow, RowFormatter

# todo Поиск по всем чатам в БД с пониманием опечаток(--search)


# todo ротация?
def _build_export_path(
    chat_id: ChatID,
    kind: str,
    ext: str,
    ts: str | None = None,
) -> Path:
    """Строит путь файла экспорта: ``exports/<chat>/<ts>_<kind>.<ext>``.

    Args:
        chat_id: ID чата (хранится "сырым"); ``0`` — полный экспорт всей БД,
            кладётся в ``exports/_full`` с именем БД в имени файла.
        kind: Назначение файла (напр. ``"filenames"``, ``"report_duplicates"``).
        ext: Расширение (напр. ``"txt"``, ``"csv"``).
        ts: Таймстемп; ``None`` — текущий момент. Для прогонов по нескольким
            чатам считается один раз и пробрасывается.

    Returns:
        Путь к файлу экспорта (каталог создаётся при необходимости).
    """
    if ts is None:
        ts = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")

    exports_root = Path(EXPORTS_DIR)
    if chat_id == 0:
        sub = exports_root / "_full"
        filename = f"{ts}_{kind}_of_{Path(DB_FILE).stem}.{ext}"
    else:
        sub = exports_root / str(chat_id)
        filename = f"{ts}_{kind}.{ext}"

    sub.mkdir(parents=True, exist_ok=True)
    return sub / filename


async def _generic_export_to_txt(
    chat_id: ChatID, output_file: str, sql_query: str, line_formatter: RowFormatter
) -> None:
    """Общая функция для экспорта данных из БД в текстовый файл.

    Args:
        chat_id: ID чата (подставляется в запрос первым параметром).
        output_file: Путь к итоговому файлу.
        sql_query: SQL-запрос с одним плейсхолдером ``?`` для chat_id.
        line_formatter: Функция ``DBRow -> str | None``; ``None`` пропускает строку.
    """
    log.info(f"Запущена задача экспорта для чата {chat_label(chat_id)} в файл '{output_file}'...")

    try:
        if not Path(DB_FILE).exists():
            log.critical(f"Файл базы данных '{DB_FILE}' не найден. Нечего экспортировать.")
            return

        async with aiosqlite.connect(DB_FILE) as conn:
            conn.row_factory = aiosqlite.Row
            async with conn.execute(sql_query, (chat_id,)) as cursor:
                data_rows = await cursor.fetchall()

        if not data_rows:
            log.warning(f"В базе данных не найдено записей для чата {chat_label(chat_id)}.")
            return

        log.info(f"Найдено {len(data_rows)} записей. Записываю в файл...")

        def write_to_file_sync():
            """(СИНХРОННАЯ!) Пишет строки экспорта в файл."""
            with open(output_file, "w", encoding="utf-8") as f:
                for row in data_rows:
                    line = line_formatter(row)
                    if line is not None:
                        f.write(line + "\n")

        await asyncio.to_thread(write_to_file_sync)
        log.info(f"Экспорт успешно завершен. Файл сохранен как '{output_file}'.")

    except aiosqlite.Error as e:
        log.critical(f"Произошла ошибка SQLite при экспорте: {e}")
    except Exception as e:
        log.critical(f"Произошла непредвиденная ошибка при экспорте: {e}", exc_info=True)


async def _generic_export_to_csv(
    chat_id: ChatID,
    kind: str,
    sql_query: str,
    sql_query_full: str,
    header: list[str],
    row_formatter: CsvRowFormatter,
) -> None:
    """Общая функция для экспорта данных из БД в CSV.

    Args:
        chat_id: ID чата; ``0`` — полный экспорт всей БД
            (используется ``sql_query_full`` без параметров).
        kind: Назначение файла (используется в имени).
        sql_query: SQL-запрос для конкретного чата (с плейсхолдером ``?``).
        sql_query_full: SQL-запрос для полного экспорта (без параметров).
        header: Заголовок CSV (список имён колонок).
        row_formatter: Функция ``DBRow -> list[str] | None``; ``None``
            пропускает запись.
    """
    is_full_export = chat_id == 0
    output_file = _build_export_path(chat_id, kind, "csv")

    if is_full_export:
        log.info(f"Запущена задача ПОЛНОГО экспорта '{kind}' в '{output_file}'...")
    else:
        log.info(
            f"Запущена задача экспорта '{kind}' для чата {chat_label(chat_id)} в '{output_file}'..."
        )

    try:
        if not Path(DB_FILE).exists():
            log.critical(f"Файл базы данных '{DB_FILE}' не найден. Нечего экспортировать.")
            return

        async with aiosqlite.connect(DB_FILE) as conn:
            conn.row_factory = aiosqlite.Row
            if is_full_export:
                query, params = sql_query_full, ()
            else:
                query, params = sql_query, (chat_id,)
            async with conn.execute(query, params) as cursor:
                data_rows = await cursor.fetchall()

        if not data_rows:
            target = "базе данных" if is_full_export else f"чате {chat_label(chat_id)}"
            log.warning(f"В {target} не найдено записей.")
            return

        log.info(f"Найдено {len(data_rows)} записей. Обработка и сохранение в CSV...")

        def write_to_file_sync():
            """(СИНХРОННАЯ!) Пишет строки экспорта в CSV."""
            import csv

            # utf-8-sig нужен, чтобы Excel автоматически правильно распознал кириллицу
            with open(output_file, "w", encoding="utf-8-sig", newline="") as f:
                # delimiter=';' используется по умолчанию в русскоязычном Excel
                writer = csv.writer(f, delimiter=";")
                writer.writerow(header)
                for row in data_rows:
                    fields = row_formatter(row)
                    if fields is not None:
                        writer.writerow(fields)

        await asyncio.to_thread(write_to_file_sync)
        log.info(f"Экспорт успешно завершен. Файл сохранен как '{output_file}'.")

    except aiosqlite.Error as e:
        log.critical(f"Произошла ошибка SQLite при экспорте: {e}")
    except Exception as e:
        log.critical(f"Произошла непредвиденная ошибка при экспорте: {e}", exc_info=True)


async def export_filenames_to_txt(chat_id: ChatID) -> None:
    """Экспортирует ТОЛЬКО имена файлов (функция-обёртка).

    Args:
        chat_id: ID чата, для которого экспортировать имена.
    """
    output_file = _build_export_path(chat_id, "filenames", "txt")

    await _generic_export_to_txt(
        chat_id=chat_id,
        output_file=str(output_file),
        sql_query="SELECT file_name FROM audios WHERE chat_id = ? ORDER BY file_name",
        line_formatter=lambda row: row["file_name"] if row["file_name"] else None,
    )


async def export_filenames_with_url_to_txt(chat_id: ChatID) -> None:
    """Экспортирует имена файлов и ссылки на сообщения (функция-обёртка).

    Args:
        chat_id: ID чата, для которого экспортировать имена со ссылками.
    """
    try:
        from wcwidth import wcswidth
    except ImportError:
        log.warning("Библиотека 'wcwidth' не найдена. Выравнивание колонок может быть неточным.")
        wcswidth = len

    if chat_id >= 0:
        log.warning("Возможно личный чат, ссылки могут быть не действительны!")
    public_chat_id = str(chat_id).removeprefix("-100")

    def formatter(row: aiosqlite.Row) -> str | None:
        if not row["file_name"]:
            return None
        file_name = row["file_name"]
        message_id = row["message_id"]
        target_width = 80  # Целевая визуальная ширина колонки

        # 1. Вычисляем реальную визуальную ширину имени файла
        visual_width = wcswidth(file_name)
        if visual_width < 0:
            visual_width = len(file_name)

        # 2. Вычисляем, сколько пробелов нужно добавить
        padding_needed = target_width - visual_width

        # 3. Если имя файла уже длиннее нашей колонки, добавим всего один пробел
        if padding_needed <= 0:
            padding_needed = 1

        padding = " " * padding_needed

        # 4. Собираем строку
        return f"{file_name}{padding}| https://t.me/c/{public_chat_id}/{message_id}"

    output_file = _build_export_path(chat_id, "filenames_with_urls", "txt")

    await _generic_export_to_txt(
        chat_id=chat_id,
        output_file=str(output_file),
        sql_query="SELECT file_name, message_id FROM audios WHERE chat_id = ? ORDER BY file_name",
        line_formatter=formatter,
    )


async def export_cleaned_names_to_csv(chat_id: ChatID) -> None:
    """Экспортирует процесс очистки имён файлов в CSV для проверки фильтров.

    Формат: Исходное имя, После ``_clean_filename``, После ``default_process``.

    Args:
        chat_id: ID чата; ``0`` — экспорт всей базы целиком.
    """

    def formatter(row: DBRow) -> list[str] | None:
        orig = row["file_name"]
        if not orig:
            return None
        cleaned = _clean_filename(orig)
        return [orig, cleaned, _process_for_fuzzy(cleaned)]

    await _generic_export_to_csv(
        chat_id=chat_id,
        kind="cleaned_names",
        sql_query="SELECT file_name FROM audios WHERE chat_id = ? AND file_name IS NOT NULL ORDER BY file_name",
        sql_query_full="SELECT file_name FROM audios WHERE file_name IS NOT NULL ORDER BY file_name",
        header=["Исходное имя", "После _clean_filename", "После default_process"],
        row_formatter=formatter,
    )


async def export_cleaned_meta_to_csv(chat_id: ChatID) -> None:
    """Экспортирует процесс очистки метаданных (performer+title) в CSV.

    Формат: Performer, Title, После ``_clean_meta``, После ``default_process``.
    Показывает ровно ту строку, которую видит fuzzy-матчер.

    Args:
        chat_id: ID чата; ``0`` — экспорт всей базы целиком.
    """

    def formatter(row: DBRow) -> list[str] | None:
        cleaned = _clean_meta(row["performer"], row["title"])
        if not cleaned:
            return None
        return [
            row["performer"] or "",
            row["title"] or "",
            cleaned,
            _process_for_fuzzy(cleaned),
        ]

    where = "(performer IS NOT NULL OR title IS NOT NULL)"
    order = "ORDER BY performer, title"
    await _generic_export_to_csv(
        chat_id=chat_id,
        kind="cleaned_meta",
        sql_query=f"SELECT performer, title FROM audios WHERE chat_id = ? AND {where} {order}",
        sql_query_full=f"SELECT performer, title FROM audios WHERE {where} {order}",
        header=["Performer", "Title", "После _clean_meta", "После default_process"],
        row_formatter=formatter,
    )


async def export_database_to_xlsx(chat_id: ChatID) -> None:
    """Универсальный экспорт БД в Excel.

    Args:
        chat_id: ID чата; ``0`` — вся база целиком, иначе данные
            фильтруются по конкретному чату.
    """
    try:
        import openpyxl
        from openpyxl.utils.cell import get_column_letter
    except ImportError:
        log.critical(
            "Для экспорта в Excel требуется библиотека openpyxl.\n"
            "Пожалуйста, установите её: pip install openpyxl"
        )
        return

    # Константы
    EXCEL_MAX_SHEET_NAME_LEN: Final[int] = 31
    EXCEL_MAX_COLUMN_WIDTH: Final[int] = 80
    EXCEL_PADDING: Final[int] = 2
    INVALID_EXCEL_CHARS: Final[frozenset[str]] = frozenset("[]:*?/\\")

    # Определяем имя файла и режим
    is_full_export = chat_id == 0
    output_file = _build_export_path(chat_id, "database_export", "xlsx")

    if is_full_export:
        log.info(f"Запущена задача ПОЛНОГО экспорта базы данных в '{output_file}'...")
    else:
        log.info(f"Запущена задача экспорта для чата {chat_id} в '{output_file}'...")

    if not Path(DB_FILE).exists():
        log.critical(f"Файл базы данных '{DB_FILE}' не найден.")
        return

    # --- ВНУТРЕННИЕ ХЕЛПЕРЫ ---

    def _sanitize_sheet_name(name: str) -> str:
        return "".join("_" if c in INVALID_EXCEL_CHARS else c for c in name)

    def _get_unique_sheet_name(base_name: str, used_names: set[str]) -> str:
        clean_name = _sanitize_sheet_name(base_name)
        safe_base = clean_name[:EXCEL_MAX_SHEET_NAME_LEN]

        if safe_base not in used_names:
            used_names.add(safe_base)
            return safe_base

        for i in range(1, 1000):
            suffix = f"_{i}"
            allowed_len = EXCEL_MAX_SHEET_NAME_LEN - len(suffix)
            candidate = clean_name[:allowed_len] + suffix
            if candidate not in used_names:
                used_names.add(candidate)
                return candidate
        return safe_base[:EXCEL_MAX_SHEET_NAME_LEN]

    def _convert_cell_value(value: Any) -> Any:
        if value is None:
            return ""
        if isinstance(value, bytes):
            return f"<BLOB {len(value)} bytes>"
        if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
            return f"'{value}"
        return value

    def _append_table_to_workbook(
        wb: openpyxl.Workbook,
        table_name: str,
        headers: list[str],
        rows: list[tuple],
        used_sheet_names: set[str],
    ):
        sheet_title = _get_unique_sheet_name(table_name, used_sheet_names)
        ws = wb.create_sheet(title=sheet_title)
        ws.append(headers)

        for row in rows:
            safe_row = [_convert_cell_value(cell) for cell in row]
            ws.append(safe_row)

        for i, col_cells in enumerate(ws.columns, start=1):
            max_len = 0
            column_letter = get_column_letter(i)
            for cell in col_cells[:50]:
                try:
                    val = cell.value
                    if val is None:
                        val_len = 0
                    elif isinstance(val, (datetime.datetime, datetime.date)):
                        val_len = 18
                    else:
                        val_len = len(str(val))
                    if val_len > max_len:
                        max_len = val_len
                except Exception:
                    continue

            adjusted_width = min(max_len + EXCEL_PADDING, EXCEL_MAX_COLUMN_WIDTH)
            ws.column_dimensions[column_letter].width = adjusted_width

    # --- ОСНОВНАЯ ЛОГИКА ---

    wb = openpyxl.Workbook()
    if wb.active:
        wb.remove(wb.active)

    used_sheet_names: set[str] = set()
    data_found = False

    try:
        async with aiosqlite.connect(DB_FILE) as conn:
            async with conn.execute("SELECT name FROM sqlite_master WHERE type='table';") as cursor:
                tables = await cursor.fetchall()
                table_names = [row[0] for row in tables]

            log.info(f"Найдено таблиц: {len(table_names)}.")

            for table in table_names:
                if table.startswith("sqlite_"):
                    continue

                try:
                    safe_table_name = table.replace('"', '""')
                    async with conn.execute(f'PRAGMA table_info("{safe_table_name}");') as cursor:
                        columns_info = await cursor.fetchall()
                        column_names = [col[1] for col in columns_info]
                except Exception as e:
                    log.error(f"Не удалось получить схему таблицы '{table}': {e}")
                    continue

                has_chat_id = "chat_id" in column_names

                if not is_full_export and has_chat_id:
                    query = f'SELECT * FROM "{safe_table_name}" WHERE chat_id = ?'
                    params = (chat_id,)
                else:
                    query = f'SELECT * FROM "{safe_table_name}"'
                    params = ()

                try:
                    async with conn.execute(query, params) as cursor:
                        db_rows = await cursor.fetchall()
                        rows_tuples = [tuple(row) for row in db_rows]

                        if not rows_tuples:
                            if is_full_export:
                                log.info(f"  -> Таблица '{table}': пуста.")
                            else:
                                log.debug(f"  -> Таблица '{table}': нет данных для этого чата.")
                            continue

                        data_found = True
                        log.info(f"  -> Таблица '{table}': экспорт {len(rows_tuples)} строк...")

                        await asyncio.to_thread(
                            _append_table_to_workbook,
                            wb,
                            table,
                            column_names,
                            rows_tuples,
                            used_sheet_names,
                        )

                except Exception as e:
                    log.error(f"Ошибка при обработке таблицы '{table}': {e}")

        if not data_found:
            log.warning("Данные для экспорта не найдены.")
            return

        log.info("Сохранение файла на диск...")
        await asyncio.to_thread(wb.save, output_file)
        log.info(f"Экспорт успешно завершен. Файл: '{output_file}'")

    except Exception as e:
        log.critical(f"Критическая ошибка при экспорте Excel: {e}", exc_info=True)
        if Path(output_file).exists():
            with suppress(OSError):
                os.remove(output_file)
